import eventlet
eventlet.monkey_patch()

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, Response
from flask_socketio import SocketIO, emit, join_room, leave_room
import json
import os
import re
from database import db, Student, Feedback, SRCUser, SystemLog, Announcement, FeedbackVote, PasswordResetToken
from database import ForumTopic, ForumReply, ForumTopicVote, ForumTopicTag, ForumReplyVote
from database import ChatRoom, ChatMessage, ChatRoomMember, ChatRoomSentiment
from database import SolutionTemplate, SolutionFeedback, CustomLexicon, UnknownWord, AIReviewLog, Notification
from database import is_valid_htu_email, extract_student_id_from_email
from sentiment_analyzer import process_feedback, analyze_chat_message, analyze_topic, get_room_sentiment_summary, get_forum_sentiment_summary, censor_text, get_sentiment_explanation, get_urgency_explanation, build_ai_explanation
from sentiment.topic_extractor import extract_topics
from solution_recommender import recommend_solutions
from enhanced_recommender import get_trending_issues, get_department_workload, get_engine
from recommender import generate_recommendation
from security_manager import SecurityManager, require_2fa, get_client_ip, get_user_agent
from recommendation_learning import RecommendationLearner
import logging
import sys

# Custom login_required decorator (flask-login not installed)
def login_required(f):
    """Decorator to require login for a route."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('student_id') and not session.get('admin_id'):
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# Configure structured logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('app.log', encoding='utf-8')
    ]
)
logger = logging.getLogger('HTU_SRC')
from logger import log_student_action, log_admin_action, log_feedback_action, log_system_action

# ==================== NOTIFICATION HELPER ====================

def create_notification(student_id, feedback_id, notification_type, title, message):
    """Create a notification for a student."""
    try:
        notification = Notification(
            recipient_type='student',
            student_id=student_id,
            feedback_id=feedback_id,
            notification_type=notification_type,
            title=title,
            message=message,
            is_read=False
        )
        db.session.add(notification)
        db.session.commit()
        return notification
    except Exception as e:
        logger.error(f"Error creating notification: {e}")
        db.session.rollback()
        return None


def create_admin_notification(feedback_id, notification_type, title, message):
    """Create a notification for admins."""
    try:
        notification = Notification(
            recipient_type='admin',
            feedback_id=feedback_id,
            notification_type=notification_type,
            title=title,
            message=message,
            is_read=False
        )
        db.session.add(notification)
        db.session.commit()
        return notification
    except Exception as e:
        logger.error(f"Error creating admin notification: {e}")
        db.session.rollback()
        return None


def notify_admins_new_feedback(feedback):
    """Notify all admins about new feedback."""
    try:
        student_name = feedback.student.full_name if feedback.student else 'Anonymous'
        title = f'New Feedback #{feedback.id}'
        message = f'{student_name} submitted feedback in {feedback.category or "General"} - Urgency {feedback.urgency_score}/5'
        
        # Get all admin users
        admins = SRCUser.query.all()
        for admin in admins:
            notification = Notification(
                recipient_type='admin',
                admin_id=admin.id,
                feedback_id=feedback.id,
                notification_type='new_feedback',
                title=title,
                message=message,
                is_read=False
            )
            db.session.add(notification)
        db.session.commit()
    except Exception as e:
        logger.error(f"Error notifying admins: {e}")
        db.session.rollback()


def check_unattended_feedback():
    """Check for unattended feedback and notify admins."""
    try:
        from datetime import timedelta
        now = datetime.utcnow()
        
        # Feedback pending for more than 24 hours
        old_pending = Feedback.query.filter(
            Feedback.status == 'Pending',
            Feedback.created_at <= now - timedelta(hours=24)
        ).all()
        
        for feedback in old_pending:
            # Check if notification already sent today
            existing = Notification.query.filter(
                Notification.feedback_id == feedback.id,
                Notification.notification_type == 'unattended',
                Notification.created_at >= now - timedelta(hours=24)
            ).first()
            
            if not existing:
                student_name = feedback.student.full_name if feedback.student else 'Anonymous'
                title = f'Unattended Feedback #{feedback.id}'
                message = f'Feedback from {student_name} has been pending for over 24 hours. Category: {feedback.category or "General"}, Urgency: {feedback.urgency_score}/5'
                
                admins = SRCUser.query.all()
                for admin in admins:
                    notification = Notification(
                        recipient_type='admin',
                        admin_id=admin.id,
                        feedback_id=feedback.id,
                        notification_type='unattended',
                        title=title,
                        message=message,
                        is_read=False
                    )
                    db.session.add(notification)
        db.session.commit()
    except Exception as e:
        logger.error(f"Error checking unattended feedback: {e}")
        db.session.rollback()

from werkzeug.security import generate_password_hash, check_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import datetime, timedelta
import secrets
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from markupsafe import Markup
from sqlalchemy import or_

# ==================== EMOTION EMOJI MAPPING ====================

EMOTION_EMOJI_MAP = {
    'anger': '😠', 'fear': '😨', 'sadness': '😢', 'joy': '😊',
    'trust': '🤝', 'disgust': '🤢', 'surprise': '😲',
    'anticipation': '🤔', 'optimism': '🌟', 'love': '❤️',
    'gratitude': '🙏', 'confusion': '😕', 'neutral': '😐',
}

EMOTION_EMOJI_MAP_COMPOUND = {
    'frustrated_resignation': '😤', 'threatened_defensive': '😰',
    'distressed_anxious': '😟', 'admiring_appreciative': '🥰',
    'delighted_surprised': '🤩', 'revolted_indignant': '🤬',
    'disheartened': '😞', 'alarmed': '😱',
    'agitated': '😤', 'anxious': '😰', 'somber': '😔',
    'cheerful': '😊', 'confident': '💪', 'repulsed': '🤢',
    'astonished': '😲', 'expectant': '🤔', 'hopeful': '🌟',
    'affectionate': '🥰', 'thankful': '🙏', 'perplexed': '😕',
}

def get_emotion_emoji(emotion_name, compound_mood=None):
    """Get emoji for an emotion, falling back to compound mood emoji."""
    if emotion_name in EMOTION_EMOJI_MAP:
        return EMOTION_EMOJI_MAP[emotion_name]
    if compound_mood and compound_mood in EMOTION_EMOJI_MAP_COMPOUND:
        return EMOTION_EMOJI_MAP_COMPOUND[compound_mood]
    return '😐'


def _parse_json_field(value, default):
    """Safely parse a JSON-encoded DB column, returning default on failure."""
    if not value:
        return default
    try:
        return json.loads(value)
    except (ValueError, TypeError):
        return default


def _get_db_templates_for_category(category):
    """Fetch admin-configured SolutionTemplate overrides for a category.

    Returns a list of dicts shaped like the hardcoded SOLUTION_TEMPLATES
    entries (plus an internal '_db_id' so the winning template can be
    credited with a usage_count bump). Returns [] when no active
    overrides exist for the category, so callers can fall back to the
    built-in templates in solution_recommender.py.
    """
    rows = SolutionTemplate.query.filter_by(category=category, is_active=True).all()
    templates = []
    for r in rows:
        keywords = [k.strip() for k in (r.keywords or '').split(',') if k.strip()]
        if not keywords:
            continue
        templates.append({
            'keywords': keywords,
            'short_term_solution': r.short_term_solution,
            'long_term_solution': r.long_term_solution,
            'responsible_department': r.responsible_department,
            'estimated_time': r.estimated_time,
            '_db_id': r.id,
        })
    return templates


def _apply_recommendation(feedback, rec):
    """Persist a Recommendation onto a Feedback row and credit the winning
    SolutionTemplate (if the recommendation came from an admin-configured
    override) with a usage_count increment."""
    feedback.recommended_keywords = ','.join(rec.matched_keywords) if rec.matched_keywords else None
    feedback.short_term_solution = rec.short_term_solution
    feedback.long_term_solution = rec.long_term_solution
    feedback.responsible_department = rec.responsible_department
    feedback.estimated_time = rec.estimated_time
    feedback.recommendation_confidence = rec.confidence
    feedback.secondary_categories = json.dumps(rec.secondary_categories) if rec.secondary_categories else None

    if rec.source_template_id and feedback.used_template_id != rec.source_template_id:
        tmpl = SolutionTemplate.query.get(rec.source_template_id)
        if tmpl:
            tmpl.usage_count = (tmpl.usage_count or 0) + 1
    feedback.used_template_id = rec.source_template_id


def _apply_enhanced_recommendation(feedback, result):
    """Persist an enhanced RecommendationResult onto a Feedback row.
    
    Stores separate student recommendation and admin action plan data.
    """
    # Store the student recommendation
    feedback.student_recommendation_summary = result.student_recommendation.summary
    feedback.student_recommendation_action = result.student_recommendation.immediate_action
    feedback.student_recommendation_contact = result.student_recommendation.who_to_contact
    feedback.student_recommendation_timeline = result.student_recommendation.expected_timeline
    
    # Store the admin action plan
    feedback.admin_action_investigation = json.dumps(result.admin_action_plan.investigation_steps)
    feedback.admin_action_corrective = json.dumps(result.admin_action_plan.corrective_actions)
    feedback.admin_action_preventive = json.dumps(result.admin_action_plan.preventive_actions)
    feedback.admin_action_department = result.admin_action_plan.responsible_department
    feedback.admin_action_priority = result.admin_action_plan.priority_level
    feedback.admin_action_escalation = result.admin_action_plan.escalation_path
    
    # Store metadata
    feedback.recommendation_sentiment_type = result.sentiment
    feedback.recommendation_urgency_level = result.urgency
    feedback.recommendation_fallback_used = result.fallback_used
    feedback.recommendation_fallback_message = result.fallback_message
    feedback.recommendation_multi_issue = result.multi_issue
    
    # Also store legacy fields for backward compatibility
    feedback.short_term_solution = result.student_recommendation.summary + " " + result.student_recommendation.immediate_action
    feedback.long_term_solution = result.admin_action_plan.corrective_actions[0] if result.admin_action_plan.corrective_actions else ""
    feedback.responsible_department = result.admin_action_plan.responsible_department
    feedback.estimated_time = result.admin_action_plan.estimated_resolution_time
    feedback.recommendation_confidence = result.confidence
    feedback.recommended_keywords = ','.join(result.all_categories[0].evidence) if result.all_categories else None
    
    # Store secondary categories
    secondary = []
    for cat in result.all_categories[1:]:
        secondary.append({
            "category": cat.name,
            "keywords": cat.evidence[:5],
            "confidence": cat.confidence,
        })
    feedback.secondary_categories = json.dumps(secondary) if secondary else None



app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or secrets.token_urlsafe(32)
if not os.environ.get('SECRET_KEY'):
    app.logger.warning('Using an auto-generated SECRET_KEY. Set the SECRET_KEY environment variable in production to preserve sessions and protect cookies.')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

# Cache-busting version for static assets (CSS/JS). Changes on every process
# start (i.e. every deploy), forcing browsers/mobile devices to fetch fresh
# files instead of serving a stale cached copy.
import time as _time
ASSET_VERSION = str(int(_time.time()))

@app.context_processor
def inject_asset_version():
    return {'asset_version': ASSET_VERSION}


def generate_csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


@app.context_processor
def inject_csrf_helpers():
    return {
        'csrf_token': generate_csrf_token,
        'csrf_field': lambda: Markup(f'<input type="hidden" name="csrf_token" value="{generate_csrf_token()}">')
    }


@app.template_filter('fromjson')
def _jinja_fromjson(value):
    """Parse a JSON-encoded DB column (e.g. Feedback.secondary_categories) in templates."""
    return _parse_json_field(value, [])


def _csrf_exempt_path():
    return request.path.startswith('/socket.io') or request.endpoint == 'static'


@app.before_request
def ensure_csrf_token():
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_urlsafe(32)


@app.before_request
def validate_csrf_token():
    if request.method not in ('POST', 'PUT', 'PATCH', 'DELETE'):
        return

    if _csrf_exempt_path():
        return

    token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
    if not token or token != session.get('_csrf_token'):
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': 'Invalid or missing CSRF token'}), 400
        return render_template('error.html', error='Invalid CSRF token. Please refresh the page and try again.'), 400

# On Render, DATABASE_URL / SQLALCHEMY_DATABASE_URI point to PostgreSQL.
# Locally (no env var), fall back to the existing SQLite file.
_DATABASE_URL = os.environ.get('DATABASE_URL') or os.environ.get('SQLALCHEMY_DATABASE_URI') or 'sqlite:///feedback.db'
if _DATABASE_URL.startswith('postgres://'):
    _DATABASE_URL = _DATABASE_URL.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
IS_SQLITE = _DATABASE_URL.startswith('sqlite')


socketio = SocketIO(app, cors_allowed_origins="*")
limiter = Limiter(get_remote_address, app=app, default_limits=["200 per day", "50 per hour"])

db.init_app(app)

# ==================== RESPONSE TEMPLATES ====================

RESPONSE_TEMPLATES = {
    'Wi-Fi Issues': "Thank you for reporting the Wi-Fi issue. Our ICT team has been notified and will investigate. Please expect an update within 48 hours.",
    'Maintenance': "Thank you for your report. Maintenance has been scheduled and will be addressed within 3 working days.",
    'Staff Conduct': "Thank you for bringing this to our attention. This matter has been escalated to the appropriate department for review.",
    'Safety Concern': "⚠️ This is a priority issue. Security has been notified and additional measures will be implemented.",
    'Accommodation': "Thank you for your feedback about accommodation. This has been forwarded to the Hall management.",
    'Catering': "Thank you for your feedback about catering. We will address this with the food services provider.",
    'General Acknowledgment': "Thank you for your feedback. We have received it and will review it with the relevant committee.",
    'Resolved Confirmation': "✅ We are pleased to inform you that this issue has been resolved. Thank you for your patience."
}

# ==================== SECURITY HEADERS ====================

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response

# ==================== GLOBAL ERROR HANDLER ====================

@app.errorhandler(500)
def internal_server_error(e):
    """Log the full traceback and return a user-friendly error page."""
    import traceback
    logger.error(f"500 Error: {e}\n{traceback.format_exc()}")
    log_system_action('Error', '500', f'Internal Server Error: {e}', 'ERROR')
    try:
        return render_template('error.html', error='Sorry, something went wrong on our end. Please try again in a moment.', error_code=500), 500
    except Exception:
        return 'Internal Server Error', 500

@app.errorhandler(404)
def not_found(e):
    logger.info(f"404 Not Found: {request.path}")
    return render_template('error.html', error='The page you are looking for was not found.', error_code=404), 404

@app.errorhandler(403)
def forbidden(e):
    logger.warning(f"403 Forbidden: {request.path} from {get_client_ip()}")
    return render_template('error.html', error='You do not have permission to access this resource.', error_code=403), 403

@app.errorhandler(429)
def rate_limited(e):
    logger.warning(f"429 Rate Limited: {request.path} from {get_client_ip()}")
    return render_template('error.html', error='Too many requests. Please try again later.', error_code=429), 429

# ==================== SESSION TIMEOUT ====================

@app.before_request
def make_session_permanent():
    session.permanent = True

@app.before_request
def check_session_timeout():
    if 'student_id' in session:
        last_activity = session.get('last_activity')
        if last_activity:
            last = datetime.fromisoformat(last_activity)
            if datetime.utcnow() - last > timedelta(hours=2):
                session.clear()
                return redirect(url_for('student_login'))
        session['last_activity'] = datetime.utcnow().isoformat()

# ==================== CREATE TABLES / SCHEMA REPAIR ====================

def _repair_chat_messages_schema_if_needed():
    """Repairs legacy SQLite schema mismatch for chat_messages.

    Old schema (found in some DBs):
      - topic_id (instead of room_id)
      - message_text (instead of message)
      - anonymous (extra)

    Current app schema expects:
      - room_id, message, cleaned_message, is_flagged
    """
    import sqlite3
    from pathlib import Path

    db_path = Path(app.root_path) / 'instance' / 'feedback.db'
    # Fallback to repo-relative path used by your config: sqlite:///feedback.db
    if not db_path.exists():
        db_path = Path('instance') / 'feedback.db'

    if not db_path.exists():
        return

    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(chat_messages)")
    cols = [r[1] for r in cur.fetchall()]

    # Already correct
    if 'room_id' in cols and 'message' in cols:
        conn.close()
        return

    # Only attempt repair when we recognize the legacy columns
    if 'topic_id' not in cols or 'message_text' not in cols:
        conn.close()
        return

    # Create new table with expected schema
    cur.execute('''
        CREATE TABLE IF NOT EXISTS chat_messages_new (
            id INTEGER PRIMARY KEY,
            room_id INTEGER NOT NULL,
            student_id VARCHAR(20) NOT NULL,
            message TEXT NOT NULL,
            cleaned_message TEXT,
            sentiment VARCHAR(10),
            sentiment_score FLOAT,
            urgency_score INTEGER DEFAULT 1,
            is_flagged BOOLEAN DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Migrate legacy data
    cur.execute('''
        INSERT OR REPLACE INTO chat_messages_new (
            id, room_id, student_id, message, cleaned_message, sentiment, sentiment_score, urgency_score, is_flagged, created_at
        )
        SELECT
            id,
            topic_id AS room_id,
            student_id,
            message_text AS message,
            NULL AS cleaned_message,
            sentiment,
            sentiment_score,
            urgency_score,
            0 AS is_flagged,
            created_at
        FROM chat_messages
    ''')

    # Swap tables
    cur.execute('DROP TABLE chat_messages')
    cur.execute('ALTER TABLE chat_messages_new RENAME TO chat_messages')

    conn.commit()
    conn.close()


def _repair_feedback_schema_if_needed():
    """Repair legacy feedback schema missing solution-recommender columns.

    Some existing feedback.db files may have older 'feedback' tables without:
      - recommended_keywords
      - short_term_solution
      - long_term_solution
      - responsible_department
      - estimated_time

    We add these columns when absent to prevent runtime 500s on /public.
    """
    import sqlite3
    from pathlib import Path

    # Match the DB location used by SQLALCHEMY_DATABASE_URI (sqlite:///feedback.db)
    db_path = Path(app.root_path) / 'instance' / 'feedback.db'
    if not db_path.exists():
        db_path = Path('instance') / 'feedback.db'
    if not db_path.exists():
        return

    conn = sqlite3.connect(str(db_path))
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(feedback)")
    cols = [r[1] for r in cur.fetchall()]
    cur.close()

    # Only attempt repair if feedback table exists (or is being created).
    # If it doesn't exist yet, db.create_all() will handle it.
    if not cols:
        conn.close()
        return

    desired_columns = {
        'recommended_keywords': 'TEXT',
        'short_term_solution': 'TEXT',
        'long_term_solution': 'TEXT',
        'responsible_department': 'TEXT',
        'estimated_time': 'TEXT',
        'confidence_score': 'FLOAT',
        'dominant_emotion': 'VARCHAR(50)',
        'compound_mood': 'VARCHAR(50)',
        'emotion_intensities': 'TEXT',
        'secondary_emotions': 'TEXT',
    }

    added_any = False
    for col, col_type in desired_columns.items():
        if col not in cols:
            conn.execute(f"ALTER TABLE feedback ADD COLUMN {col} {col_type}")
            added_any = True

    conn.commit()
    conn.close()


# ==================== NLTK DATA SAFETY NET ====================

def _ensure_schema_aligned():
    """Ensure every model table has all columns required by its SQLAlchemy model.

    db.create_all() only creates MISSING tables; it does NOT add new columns to
    EXISTING tables. On the deployed PostgreSQL database (Render), tables may
    have been originally created by an older release and be missing newer columns
    (e.g. the solution-recommender and confidence/emotion columns on `feedback`).
    Without them, INSERT/UPDATE statements fail with a 500 Internal Server Error.

This function is dialect-agnostic (works on both SQLite and PostgreSQL):
      - Uses SQLAlchemy's inspector to read existing columns per table.
      - Adds any missing columns via ALTER TABLE ... ADD COLUMN.

    IMPORTANT: PostgreSQL does NOT allow ALTER TABLE ... ADD COLUMN with a
    NOT NULL constraint unless a DEFAULT is provided. All our model columns
    that are nullable=False already have a client-side Python default, but the
    compiled DDL may still include NOT NULL. To be safe for the deployment
    migration, we add columns as NULLABLE (omitting NOT NULL) here. The
    application always supplies values on INSERT, so NULLs are never observed
    in practice, and this avoids PostgreSQL migration failures.
    """
    import traceback
    from sqlalchemy import inspect as sa_inspect, text as sa_text

    insp = sa_inspect(db.engine)
    existing_tables = set(insp.get_table_names())

    added_any = False
    try:
        for table in db.metadata.sorted_tables:
            table_name = table.name
            if table_name not in existing_tables:
                continue  # db.create_all() will create it.

            existing_cols = {col['name'] for col in insp.get_columns(table_name)}

            for col in table.columns:
                if col.name in existing_cols:
                    continue

                # Build the base portable SQL type string for the ALTER statement.
                col_type = col.type
                try:
                    compiled_type = col_type.compile(dialect=db.engine.dialect)
                except Exception:
                    compiled_type = str(col_type)

                # Strip any trailing NOT NULL / UNIQUE from the compiled type so
                # PostgreSQL does not reject adding a column without a default.
                cleaned_type = compiled_type.split(" NOT NULL")[0].strip()
                cleaned_type = cleaned_type.split(" UNIQUE")[0].strip()

                sql = f'ALTER TABLE {table_name} ADD COLUMN {col.name} {cleaned_type}'
                try:
                    db.session.execute(sa_text(sql))
                    added_any = True
                    print(f"[Schema] Added column '{table_name}.{col.name}' ({cleaned_type})")
                except Exception as e:
                    # Column may have been added concurrently / already exists.
                    print(f"[Schema] Could not add column '{table_name}.{col.name}': {e}")
                    db.session.rollback()
        if added_any:
            db.session.commit()
            print("[Schema] Added missing columns to the database.")
    except Exception as e:
        db.session.rollback()
        print(f"[Schema] ERROR during schema alignment: {e}")
        traceback.print_exc()


def _ensure_nltk_data():
    """Ensure required NLTK corpora are available at runtime.

    If the build step (download_nltk_data.py) was skipped or failed,
    this provides a fallback download at first app startup.

    To avoid blocking every worker start with a slow/blocking network
    download (which on Render can cause worker timeouts / 500s), we:
      - Search *all* NLTK data paths (not just the first), because the
        actual corpus may live in a non-default path (e.g. %APPDATA%/nltk_data).
      - Apply a short socket timeout so offline/slow startups fail fast
        instead of hanging the whole app for minutes.
      - Only attempt a download for resources that are genuinely missing,
        and never retry a resource that we already tried (cached in a
        module-level set). Each resource is checked once and cached so this
        function is cheap on subsequent calls.
    """
    import nltk

    nltk_resources = [
        'wordnet', 'punkt', 'averaged_perceptron_tagger', 'omw-1.4',
        'sentiwordnet',
    ]

    # Cache the availability result so we don't repeatedly attempt network
    # I/O on every startup / every request (which caused 13s+ delays and
    # gunicorn worker timeouts on Render).
    if getattr(_ensure_nltk_data, '_checked', False):
        return
    _ensure_nltk_data._checked = True

    # NOTE: Do NOT use socket.setdefaulttimeout() here. That mutates the process-
    # wide default and races with eventlet's listening socket creation: the
    # listener inherits the short timeout and crashes accept() with TimeoutError
    # after a few idle seconds. Downloads run in a daemon thread and are already
    # wrapped in try/except so a slow/unreachable NLTK host cannot take down the app.

    # NLTK places resources under different subdirectories depending on type:
    #   - corpora:  wordnet, sentiwordnet, omw-1.4, ...
    #   - tokenizers: punkt
    #   - taggers:  averaged_perceptron_tagger
    _NLTK_SUBDIRS = ['corpora', 'tokenizers', 'taggers']

    def _resource_available(resource):
        """Check whether a resource exists in ANY configured NLTK data path.

        NLTK groups resources by type (corpora/tokenizers/taggers), so we
        search each possible subdirectory for both unzipped and zipped forms.
        """
        from pathlib import Path
        for base in nltk.data.path:
            for sub in _NLTK_SUBDIRS:
                candidate = Path(base) / sub / resource
                if candidate.exists():
                    return True
                # Support zipped resources (resource.zip)
                if candidate.with_suffix('.zip').exists():
                    return True
        return False

    for resource in nltk_resources:
        try:
            if _resource_available(resource):
                continue
        except Exception:
            # Fall back to nltk's own lookup; if it fails we'll try download.
            try:
                nltk.data.find(f'corpora/{resource}')
                continue
            except LookupError:
                pass

        # Resource is genuinely missing. Attempt a single download.
        # Never raise — a missing corpus should degrade the sentiment
        # pipeline gracefully, not crash the app.
        try:
            nltk.download(resource, quiet=True)
            print(f"[NLTK] Downloaded missing resource: {resource}")
        except Exception as e:
            print(f"[NLTK] WARNING: Could not download {resource}: {e}")



with app.app_context():
    # Run NLTK data downloads in a background daemon thread so they never
    # block (or crash) web-worker startup. The sentiment pipeline already
    # degrades gracefully when NLTK corpora are unavailable, so a slow or
    # unreachable NLTK server must not prevent the app from booting.
    import threading as _threading
    _nltk_thread = _threading.Thread(target=_ensure_nltk_data, daemon=True)
    _nltk_thread.start()

    # SQLite-only legacy schema repairs (skip on PostgreSQL/Render).
    if IS_SQLITE:
        _repair_chat_messages_schema_if_needed()
        _repair_feedback_schema_if_needed()
    db.create_all()
    # Dialect-agnostic: ensure every model table has all columns the model
    # expects. Critical on PostgreSQL (Render) where db.create_all() does NOT
    # add missing columns to an existing table.
    _ensure_schema_aligned()
    log_system_action('Database', 'Initialization', 'Database tables created')
    
    if not SRCUser.query.filter_by(username='admin').first():
        admin = SRCUser(
            username='admin',
            password_hash=generate_password_hash('admin123'),
            full_name='SRC Administrator',
            role='President'
        )
        db.session.add(admin)
        db.session.commit()
        log_system_action('Admin', 'Creation', 'Default admin account created')


def add_db_log(log_type, level, user_type, user_id, action, details):
    try:
        log = SystemLog(
            log_type=log_type, level=level, user_type=user_type,
            user_id=str(user_id)[:50], action=action[:100], details=details[:500],
            ip_address=request.remote_addr if request else '0.0.0.0'
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        print(f"Failed to add DB log: {e}")

# ==================== AUTHENTICATION MIDDLEWARE ====================

def student_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'student_id' not in session:
            return redirect(url_for('student_login'))
        return f(*args, **kwargs)
    return decorated_function

def src_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== STUDENT AUTHENTICATION ====================

@app.route('/student/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")

def student_login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password')
        
        if not is_valid_htu_email(email):
            return render_template('student_login.html', error='Invalid email format. Use: studentid@htu.edu.gh')
        
        student_id = extract_student_id_from_email(email)
        student = Student.query.filter_by(student_id=student_id).first()
        
        if student and check_password_hash(student.password_hash, password):
            session['student_id'] = student.student_id
            session['student_name'] = student.full_name
            session['student_email'] = student.email
            student.last_login = datetime.utcnow()
            db.session.commit()
            
            log_student_action(student.student_id, 'Login Success', 'Logged in')
            add_db_log('auth', 'INFO', 'student', student.student_id, 'Login Success', '')
            return redirect(url_for('student_dashboard'))
        else:
            log_student_action(student_id or 'unknown', 'Login Failed', 'Invalid password', 'WARNING')
            return render_template('student_login.html', error='Invalid email or password')
    
    return render_template('student_login.html')

@app.route('/student/register', methods=['GET', 'POST'])
@limiter.limit("3 per hour")
def student_register():
    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip()
        email = request.form.get('email', '').strip().lower()
        full_name = request.form.get('full_name', '').strip()
        department = request.form.get('department', '')
        year_of_study = request.form.get('year_of_study', 1, type=int)
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        expected_email = f"{student_id}@htu.edu.gh"
        if email != expected_email:
            return render_template('student_register.html', error=f'Email must be: {expected_email}')
        
        if Student.query.filter_by(student_id=student_id).first():
            return render_template('student_register.html', error='Student ID already registered')
        
        if Student.query.filter_by(email=email).first():
            return render_template('student_register.html', error='Email already registered')
        
        if password != confirm_password:
            return render_template('student_register.html', error='Passwords do not match')
        
        if len(password) < 6:
            return render_template('student_register.html', error='Password must be at least 6 characters')
        
        student = Student(
            student_id=student_id, email=email, password_hash=generate_password_hash(password),
            full_name=full_name, department=department, year_of_study=year_of_study
        )
        db.session.add(student)
        db.session.commit()
        
        log_student_action(student_id, 'Registration Success', f'Student {full_name} registered')
        
        session['student_id'] = student.student_id
        session['student_name'] = student.full_name
        session['student_email'] = student.email
        
        return redirect(url_for('student_dashboard'))
    
    return render_template('student_register.html')

@app.route('/student/forgot-password', methods=['GET', 'POST'])
@limiter.limit("3 per hour")
def forgot_password():
    if request.method == 'POST':
        student_id = request.form.get('student_id', '').strip()
        student = Student.query.filter_by(student_id=student_id).first()
        
        if student:
            token = secrets.token_urlsafe(32)
            expires = datetime.utcnow() + timedelta(hours=1)
            PasswordResetToken.query.filter_by(student_id=student.student_id).delete()
            reset = PasswordResetToken(student_id=student.student_id, token=token, expires_at=expires)
            db.session.add(reset)
            db.session.commit()
            
            log_student_action(student.student_id, 'Password Reset Request', 'Reset token generated')
            return render_template('forgot_password.html', token=token, student_id=student.student_id, message='Your reset token has been generated.')
        
        return render_template('forgot_password.html', error='Student ID not found')
    
    return render_template('forgot_password.html')

@app.route('/student/reset-password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        token = request.form.get('token')
        password = request.form.get('password')
        confirm = request.form.get('confirm_password')
        
        reset = PasswordResetToken.query.filter_by(student_id=student_id, token=token).first()
        
        if not reset or reset.expires_at < datetime.utcnow():
            return render_template('reset_password.html', error='Invalid or expired reset token')
        
        if password != confirm:
            return render_template('reset_password.html', error='Passwords do not match', token=token, student_id=student_id)
        
        if len(password) < 6:
            return render_template('reset_password.html', error='Password must be at least 6 characters', token=token, student_id=student_id)
        
        student = Student.query.filter_by(student_id=student_id).first()
        student.password_hash = generate_password_hash(password)
        db.session.delete(reset)
        db.session.commit()
        
        log_student_action(student.student_id, 'Password Reset', 'Password changed successfully')
        return redirect(url_for('student_login'))
    
    return render_template('reset_password.html')

@app.route('/student/logout', methods=['POST'])
def student_logout():
    student_id = session.get('student_id')
    log_student_action(student_id, 'Logout', 'User logged out')
    session.clear()
    return redirect(url_for('index'))

@app.route('/student/change-password', methods=['GET', 'POST'])
@student_required
@limiter.limit("5 per hour")
def student_change_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        student = Student.query.filter_by(student_id=session['student_id']).first()

        if not student or not check_password_hash(student.password_hash, current_password):
            return render_template('change_password.html', role='student', error='Current password is incorrect')

        if new_password != confirm_password:
            return render_template('change_password.html', role='student', error='New passwords do not match')

        if len(new_password) < 6:
            return render_template('change_password.html', role='student', error='New password must be at least 6 characters')

        if check_password_hash(student.password_hash, new_password):
            return render_template('change_password.html', role='student', error='New password must be different from your current password')

        student.password_hash = generate_password_hash(new_password)
        db.session.commit()

        log_student_action(student.student_id, 'Password Change', 'Password changed successfully from dashboard')
        add_db_log('auth', 'INFO', 'student', student.student_id, 'Password Change', '')
        return render_template('change_password.html', role='student', success='Your password has been updated successfully.')

    return render_template('change_password.html', role='student')

# ==================== STUDENT DASHBOARD & FEEDBACK ====================


def _dashboard_topic_summary(feedbacks, limit=5):
    """Return the most common topics in a feedback collection."""
    counts = {}
    for f in feedbacks:
        text = (getattr(f, 'cleaned_text', None) or getattr(f, 'feedback_text', None) or '').strip()
        for topic in extract_topics(text):
            counts[topic] = counts.get(topic, 0) + 1
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:limit]


def _duplicate_feedback_ids(feedbacks, threshold=0.72):
    """Find near-duplicate feedback without a new database dependency.

    Uses token Jaccard similarity on recent text. This is an early-warning
    signal for repeated reports, not a replacement for exact duplicate checks.
    """
    items = []
    for f in feedbacks:
        text = (getattr(f, 'cleaned_text', None) or getattr(f, 'feedback_text', None) or '').lower()
        tokens = {t for t in re.findall(r"[a-z0-9']+", text) if len(t) > 2}
        if tokens:
            items.append((f.id, tokens))
    duplicate_ids = set()
    for i, (fid, a) in enumerate(items):
        for gid, b in items[i+1:]:
            score = len(a & b) / max(1, len(a | b))
            if score >= threshold:
                duplicate_ids.update((fid, gid))
    return duplicate_ids


@app.route('/student/dashboard')
@student_required
def student_dashboard():
    student = Student.query.filter_by(student_id=session['student_id']).first()
    my_feedback = Feedback.query.filter_by(student_id=session['student_id']).order_by(Feedback.created_at.desc()).all()

    # Build per-feedback explanations for the UI (no DB schema change).
    # We use cleaned_text when available; fallback to feedback_text.
    for f in my_feedback:
        base_text = (f.cleaned_text or '').strip() or (f.feedback_text or '').strip()
        f.sentiment_explanation = get_sentiment_explanation(base_text).get('sentiment_explanation', [])
        f.urgency_explanation = get_urgency_explanation(base_text, f.sentiment)
        # Attach emotion & confidence display helpers
        f.confidence_display = f.confidence_score
        f.emotion_emoji = get_emotion_emoji(f.dominant_emotion, f.compound_mood)
        f.emotion_intensities_list = _parse_json_field(f.emotion_intensities, {})
        f.secondary_emotions_list = _parse_json_field(f.secondary_emotions, [])
        f.ai_topics = extract_topics(base_text)
        f.ai_explanation = build_ai_explanation(
            base_text,
            category=f.category,
            recommendation={
                'short_term_solution': f.short_term_solution,
            },
            final_sentiment=f.sentiment,
            final_confidence=f.confidence_score,
        )

    stats = {

        'total': len(my_feedback),
        'resolved': sum(1 for f in my_feedback if f.status == 'Resolved'),
        'in_progress': sum(1 for f in my_feedback if f.status == 'In Progress'),
        'pending': sum(1 for f in my_feedback if f.status == 'Pending')
    }

    student_topics = _dashboard_topic_summary(my_feedback)
    return render_template('student_dashboard.html', feedbacks=my_feedback, stats=stats,
                           student=student, student_topics=student_topics)


@app.route('/submit', methods=['GET', 'POST'])
@student_required
@limiter.limit("5 per minute")




def submit_feedback():

    if request.method == 'POST':
        text = request.form.get('feedback_text')
        category = request.form.get('category')
        location = request.form.get('location')
        anonymous = request.form.get('anonymous') == 'on'
        user_urgency = int(request.form.get('user_urgency', 3))
        
        analysis = process_feedback(text, category)
        
        # Cap user-set urgency based on sentiment
        sentiment = analysis.get('sentiment', 'Neutral')
        max_allowed_urgency = 5
        if sentiment == 'Positive':
            max_allowed_urgency = 2
        elif sentiment == 'Neutral':
            max_allowed_urgency = 3
        
        user_urgency = min(user_urgency, max_allowed_urgency)
        final_urgency = max(user_urgency, analysis['urgency_score'])
        final_urgency = min(final_urgency, max_allowed_urgency)
        
        detected_category = analysis['detected_category'] if category == 'Other' else category
        # Enhanced recommendation with sentiment, emotion, and confidence data
        emotion_data = analysis.get('emotion', None)
        sentiment_label = analysis.get('sentiment', None)
        sentiment_val = analysis.get('sentiment_score', None)

        # Use the new enhanced recommender with fallback to old recommender
        try:
            rec_result = generate_recommendation(
                text=text,
                category=detected_category,
                urgency_score=final_urgency,
                sentiment=sentiment_label,
                sentiment_score=sentiment_val,
                emotion=emotion_data,
            )
        except Exception as e:
            logger.error(f"Enhanced recommender failed: {e}, falling back to legacy recommender")
            rec = recommend_solutions(
                text=text,
                category=detected_category,
                urgency_score=final_urgency,
                sentiment=sentiment_label,
                sentiment_score=sentiment_val,
                emotion=emotion_data,
                db_templates=_get_db_templates_for_category(detected_category),
            )
            # Convert legacy recommendation to enhanced format
            from recommender import RecommendationResult, StudentRecommendation, AdminActionPlan, CategoryMatch
            rec_result = RecommendationResult(
                primary_category=detected_category,
                all_categories=[CategoryMatch(name=detected_category, confidence=rec.confidence, evidence=rec.matched_keywords)],
                sentiment=sentiment_label or "neutral",
                urgency="medium",
                student_recommendation=StudentRecommendation(
                    summary=rec.short_term_solution or "",
                    immediate_action="",
                    who_to_contact=rec.responsible_department or "SRC Secretariat",
                    expected_timeline=rec.estimated_time or "3-7 days",
                ),
                admin_action_plan=AdminActionPlan(
                    investigation_steps=[],
                    corrective_actions=[rec.long_term_solution] if rec.long_term_solution else [],
                    preventive_actions=[],
                    responsible_department=rec.responsible_department or "SRC Secretariat",
                    priority_level="medium",
                    estimated_resolution_time=rec.estimated_time or "3-7 days",
                    escalation_path="SRC Secretariat",
                ),
                confidence=rec.confidence,
            )

        # Persist confidence score and emotion data from hybrid analysis
        emotion_data = analysis.get('emotion', {})
        confidence_val = analysis.get('confidence', None)

        feedback = Feedback(
            student_id=session['student_id'],
            anonymous=anonymous,
            category=detected_category,
            location=location,
            feedback_text=text,
            cleaned_text=analysis['cleaned_text'],
            sentiment=analysis['sentiment'],
            sentiment_score=analysis['sentiment_score'],
            urgency_score=final_urgency,
            has_profanity=analysis['has_profanity'],
            # New confidence & emotion fields
            confidence_score=confidence_val,
            dominant_emotion=emotion_data.get('dominant_emotion') if emotion_data else None,
            compound_mood=emotion_data.get('compound_mood') if emotion_data else None,
            emotion_intensities=json.dumps(emotion_data.get('emotion_intensities')) if emotion_data and emotion_data.get('emotion_intensities') else None,
            secondary_emotions=json.dumps(emotion_data.get('secondary_emotions')) if emotion_data and emotion_data.get('secondary_emotions') else None,
        )
        _apply_enhanced_recommendation(feedback, rec_result)

        db.session.add(feedback)
        db.session.commit()
        
        # Notify admins about new feedback
        notify_admins_new_feedback(feedback)
        
        if analysis['has_profanity']:
            log_feedback_action(feedback.id, session['student_id'], 'Profanity Detected', 
                               'Feedback contained inappropriate language', 'WARNING')
        
        log_feedback_action(feedback.id, session['student_id'], 'Submit', 
                           f'Category: {feedback.category}, Urgency: {final_urgency}')
        
        return render_template('submit_success.html', feedback_id=feedback.id)
    
    return render_template('submit_feedback.html')

@app.route('/edit-feedback/<int:feedback_id>', methods=['GET', 'POST'])
@student_required
def edit_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    
    if feedback.student_id != session['student_id']:
        return render_template('error.html', error="Unauthorized"), 403
    
    time_limit = feedback.created_at + timedelta(hours=1)
    if datetime.utcnow() > time_limit:
        return render_template('error.html', error="Feedback can only be edited within 1 hour of submission")
    
    if request.method == 'POST':
        new_text = request.form.get('feedback_text')
        feedback.feedback_text = new_text
        feedback.category = request.form.get('category')
        feedback.location = request.form.get('location')
        feedback.urgency_score = int(request.form.get('user_urgency', 3))
        
        analysis = process_feedback(new_text, feedback.category)
        emotion_data = analysis.get('emotion', {})
        feedback.cleaned_text = analysis['cleaned_text']
        feedback.sentiment = analysis['sentiment']
        feedback.sentiment_score = analysis['sentiment_score']
        feedback.urgency_score = max(feedback.urgency_score, analysis['urgency_score'])
        feedback.has_profanity = analysis['has_profanity']
        feedback.confidence_score = analysis.get('confidence', None)
        feedback.dominant_emotion = emotion_data.get('dominant_emotion') if emotion_data else None
        feedback.compound_mood = emotion_data.get('compound_mood') if emotion_data else None
        feedback.emotion_intensities = json.dumps(emotion_data.get('emotion_intensities')) if emotion_data and emotion_data.get('emotion_intensities') else None
        feedback.secondary_emotions = json.dumps(emotion_data.get('secondary_emotions')) if emotion_data and emotion_data.get('secondary_emotions') else None

        # Recompute the solution recommendation too -- the category or
        # wording may have changed enough to point at a different fix.
        rec_result = generate_recommendation(
            text=new_text,
            category=feedback.category,
            urgency_score=feedback.urgency_score,
            sentiment=feedback.sentiment,
            sentiment_score=feedback.sentiment_score,
            emotion=emotion_data,
        )
        _apply_enhanced_recommendation(feedback, rec_result)

        db.session.commit()
        log_feedback_action(feedback.id, session['student_id'], 'Edit', 'Feedback edited')
        return redirect(url_for('student_dashboard'))
    
    return render_template('edit_feedback.html', feedback=feedback)

@app.route('/delete-feedback/<int:feedback_id>', methods=['POST'])
@student_required
def delete_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    
    if feedback.student_id != session['student_id']:
        return render_template('error.html', error="Unauthorized"), 403
    
    if feedback.status != 'Pending' and datetime.utcnow() > feedback.created_at + timedelta(hours=1):
        return render_template('error.html', error="Feedback can only be deleted if pending or within 1 hour")
    
    db.session.delete(feedback)
    db.session.commit()
    log_feedback_action(feedback_id, session['student_id'], 'Delete', 'Feedback deleted')
    return redirect(url_for('student_dashboard'))

@app.route('/vote/<int:feedback_id>', methods=['POST'])
@student_required
def vote_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    
    if feedback.student_id == session['student_id']:
        return jsonify({'error': 'Cannot vote on your own feedback'}), 400
    
    existing = FeedbackVote.query.filter_by(feedback_id=feedback_id, student_id=session['student_id']).first()
    
    if existing:
        db.session.delete(existing)
        action = 'removed'
    else:
        vote = FeedbackVote(feedback_id=feedback_id, student_id=session['student_id'])
        db.session.add(vote)
        action = 'added'
    
    db.session.commit()
    return jsonify({'success': True, 'action': action, 'vote_count': feedback.vote_count})

# ==================== FORUM ROUTES ====================

@app.route('/forum')
@student_required
def forum_index():
    category = request.args.get('category', '')
    sort = request.args.get('sort', 'latest')
    sentiment_filter = request.args.get('sentiment', '')
    
    query = ForumTopic.query
    if category:
        query = query.filter_by(category=category)
    if sentiment_filter:
        query = query.filter_by(sentiment=sentiment_filter)
    
    if sort == 'latest':
        topics = query.order_by(ForumTopic.created_at.desc()).all()
    elif sort == 'hottest':
        topics = query.all()
        topics.sort(key=lambda t: (len(t.replies) + t.urgency_score * 2), reverse=True)
    elif sort == 'most_voted':
        topics = query.all()
        topics.sort(key=lambda t: t.vote_count, reverse=True)
    else:
        topics = query.all()
        topics.sort(key=lambda t: t.vote_count, reverse=True)

    
    pinned_topics = [t for t in topics if t.is_pinned]
    regular_topics = [t for t in topics if not t.is_pinned]
    
    forum_summary = get_forum_sentiment_summary(ForumTopic.query.all())
    
    category_counts = {}
    for t in ForumTopic.query.all():
        category_counts[t.category] = category_counts.get(t.category, 0) + 1
    
    return render_template('forum_index.html', pinned_topics=pinned_topics, topics=regular_topics,
                         forum_summary=forum_summary, category_counts=category_counts,
                         current_category=category, current_sort=sort, current_sentiment=sentiment_filter)

@app.route('/forum/topic/<int:topic_id>', methods=['GET', 'POST'])
@student_required
def view_topic(topic_id):
    topic = ForumTopic.query.get_or_404(topic_id)
    student_id = session['student_id']
    
    topic.view_count += 1
    db.session.commit()
    
    user_vote = ForumTopicVote.query.filter_by(topic_id=topic_id, student_id=student_id).first()
    
    if request.method == 'POST':
        if topic.is_locked:
            return render_template('forum_topic.html', topic=topic, replies=topic.replies, 
                                 user_vote=user_vote, error='This topic is locked.')
        
        content = request.form.get('content')
        if content:
            analysis = analyze_chat_message(content)
            reply = ForumReply(
                topic_id=topic_id, student_id=student_id, content=content,
                cleaned_content=analysis['cleaned_message'], sentiment=analysis['sentiment'],
                sentiment_score=analysis['sentiment_score'], urgency_score=analysis['urgency_score'],
                is_flagged=analysis['is_flagged']
            )
            db.session.add(reply)
            topic.updated_at = datetime.utcnow()
            
            topic_analysis = analyze_topic(topic.content, topic.replies + [reply])
            topic.sentiment = topic_analysis['sentiment']
            topic.sentiment_score = topic_analysis['sentiment_score']
            topic.urgency_score = topic_analysis['urgency_score']
            
            db.session.commit()
            log_student_action(student_id, 'Forum Reply', f'Replied to topic: {topic.title}')
            return redirect(url_for('view_topic', topic_id=topic_id))
    
    replies = ForumReply.query.filter_by(topic_id=topic_id).order_by(ForumReply.created_at.asc()).all()

    reply_votes = ForumReplyVote.query.filter_by(student_id=student_id).filter(
        ForumReplyVote.reply_id.in_([r.id for r in replies])
    ).all() if replies else []
    reply_user_votes = {v.reply_id: v.vote_type for v in reply_votes}

    return render_template('forum_topic.html', topic=topic, replies=replies, user_vote=user_vote,
                            student_id=student_id, reply_user_votes=reply_user_votes)

@app.route('/forum/create', methods=['GET', 'POST'])
@student_required
def create_topic():
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        category = request.form.get('category')
        tags = request.form.get('tags', '').split(',')
        
        analysis = analyze_topic(content, [])
        
        topic = ForumTopic(
            title=title, content=content, category=category, student_id=session['student_id'],
            sentiment=analysis['sentiment'], sentiment_score=analysis['sentiment_score'],
            urgency_score=analysis['urgency_score']
        )
        db.session.add(topic)
        db.session.commit()
        
        for tag in tags:
            tag = tag.strip().lower()
            if tag:
                topic_tag = ForumTopicTag(topic_id=topic.id, tag=tag)
                db.session.add(topic_tag)
        
        db.session.commit()
        log_student_action(session['student_id'], 'Create Topic', f'Created topic: {title}')
        return redirect(url_for('view_topic', topic_id=topic.id))
    
    return render_template('create_topic.html')

@app.route('/forum/topic/<int:topic_id>/vote', methods=['POST'])
@student_required
def vote_topic(topic_id):
    vote_type = request.json.get('vote_type')
    if vote_type not in ('up', 'down'):
        return jsonify({'error': 'Invalid vote type'}), 400

    student_id = session['student_id']
    
    existing = ForumTopicVote.query.filter_by(topic_id=topic_id, student_id=student_id).first()
    
    if existing:
        if existing.vote_type == vote_type:
            db.session.delete(existing)
            action = 'removed'
        else:
            existing.vote_type = vote_type
            action = 'changed'
    else:
        vote = ForumTopicVote(topic_id=topic_id, student_id=student_id, vote_type=vote_type)
        db.session.add(vote)
        action = 'added'
    
    db.session.commit()
    
    upvotes = ForumTopicVote.query.filter_by(topic_id=topic_id, vote_type='up').count()
    downvotes = ForumTopicVote.query.filter_by(topic_id=topic_id, vote_type='down').count()
    
    return jsonify({'success': True, 'action': action, 'upvotes': upvotes, 'downvotes': downvotes, 'net_votes': upvotes - downvotes})

@app.route('/forum/reply/<int:reply_id>/vote', methods=['POST'])
@student_required
def vote_reply(reply_id):
    reply = ForumReply.query.get_or_404(reply_id)
    vote_type = (request.json or {}).get('vote_type')
    if vote_type not in ('up', 'down'):
        return jsonify({'error': 'Invalid vote type'}), 400

    student_id = session['student_id']

    if reply.student_id == student_id:
        return jsonify({'error': 'Cannot vote on your own reply'}), 400

    existing = ForumReplyVote.query.filter_by(reply_id=reply_id, student_id=student_id).first()

    if existing:
        if existing.vote_type == vote_type:
            db.session.delete(existing)
            action = 'removed'
        else:
            existing.vote_type = vote_type
            action = 'changed'
    else:
        vote = ForumReplyVote(reply_id=reply_id, student_id=student_id, vote_type=vote_type)
        db.session.add(vote)
        action = 'added'

    db.session.commit()

    upvotes = ForumReplyVote.query.filter_by(reply_id=reply_id, vote_type='up').count()
    downvotes = ForumReplyVote.query.filter_by(reply_id=reply_id, vote_type='down').count()

    return jsonify({'success': True, 'action': action, 'upvotes': upvotes, 'downvotes': downvotes, 'net_votes': upvotes - downvotes})

@app.route('/forum/reply/<int:reply_id>/edit', methods=['POST'])
@student_required
def edit_reply(reply_id):
    reply = ForumReply.query.get_or_404(reply_id)
    
    if reply.student_id != session['student_id']:
        return jsonify({'error': 'Unauthorized'}), 403
    
    time_diff = datetime.utcnow() - reply.created_at
    if time_diff.total_seconds() > 300:
        return jsonify({'error': 'Replies can only be edited within 5 minutes'}), 400
    
    new_content = request.json.get('content')
    if new_content:
        reply.content = new_content
        reply.is_edited = True
        reply.edited_at = datetime.utcnow()
        
        analysis = analyze_chat_message(new_content)
        reply.cleaned_content = analysis['cleaned_message']
        reply.sentiment = analysis['sentiment']
        reply.sentiment_score = analysis['sentiment_score']
        reply.urgency_score = analysis['urgency_score']
        reply.is_flagged = analysis['is_flagged']
        
        topic = ForumTopic.query.get(reply.topic_id)
        topic_analysis = analyze_topic(topic.content, topic.replies)
        topic.sentiment = topic_analysis['sentiment']
        topic.sentiment_score = topic_analysis['sentiment_score']
        topic.urgency_score = topic_analysis['urgency_score']
        
        db.session.commit()
        return jsonify({'success': True})
    
    return jsonify({'error': 'No content provided'}), 400

@app.route('/forum/topic/<int:topic_id>/pin', methods=['POST'])
@src_required
def pin_topic(topic_id):
    topic = ForumTopic.query.get_or_404(topic_id)
    topic.is_pinned = not topic.is_pinned
    db.session.commit()
    log_admin_action(session['admin_name'], 'Pin Topic', f'Pinned: {topic.title}')
    return jsonify({'success': True, 'is_pinned': topic.is_pinned})

@app.route('/forum/topic/<int:topic_id>/lock', methods=['POST'])
@src_required
def lock_topic(topic_id):
    topic = ForumTopic.query.get_or_404(topic_id)
    topic.is_locked = not topic.is_locked
    db.session.commit()
    log_admin_action(session['admin_name'], 'Lock Topic', f'Locked: {topic.title}')
    return jsonify({'success': True, 'is_locked': topic.is_locked})

@app.route('/forum/topic/<int:topic_id>/delete', methods=['POST'])
@src_required
def delete_topic(topic_id):
    topic = ForumTopic.query.get_or_404(topic_id)
    ForumReply.query.filter_by(topic_id=topic_id).delete()
    ForumTopicVote.query.filter_by(topic_id=topic_id).delete()
    ForumTopicTag.query.filter_by(topic_id=topic_id).delete()
    db.session.delete(topic)
    db.session.commit()
    log_admin_action(session['admin_name'], 'Delete Topic', f'Deleted topic ID: {topic_id}')
    return jsonify({'success': True})

@app.route('/forum/search')
@student_required
def search_topics():
    query = request.args.get('q', '')
    if not query:
        return redirect(url_for('forum_index'))
    
    topics = ForumTopic.query.filter(
        ForumTopic.title.contains(query) | ForumTopic.content.contains(query)
    ).order_by(ForumTopic.created_at.desc()).all()
    
    return render_template('forum_search.html', topics=topics, query=query)

@app.route('/forum/my-topics')
@student_required
def my_topics():
    topics = ForumTopic.query.filter_by(student_id=session['student_id']).order_by(ForumTopic.created_at.desc()).all()
    return render_template('forum_my_topics.html', topics=topics)

# ==================== CHAT ROUTES ====================

@app.route('/chat')
@student_required
def chat_index():
    student_id = session['student_id']
    all_rooms = ChatRoom.query.filter_by(is_active=True).order_by(ChatRoom.last_activity.desc()).all()
    my_room_ids = [m.room_id for m in ChatRoomMember.query.filter_by(student_id=student_id).all()]
    
    rooms_with_info = []
    for room in all_rooms:
        sentiment_summary = get_room_sentiment_summary(room.messages)
        rooms_with_info.append({
            'room': room,
            'is_member': room.id in my_room_ids,
            'sentiment': sentiment_summary,
            'member_count': len(room.members),
            'message_count': len(room.messages)
        })
    
    return render_template('chat_index.html', rooms=rooms_with_info)

@app.route('/chat/room/<int:room_id>')
@student_required
def chat_room(room_id):
    room = ChatRoom.query.get_or_404(room_id)
    student_id = session['student_id']
    
    existing = ChatRoomMember.query.filter_by(room_id=room_id, student_id=student_id).first()
    if not existing:
        member = ChatRoomMember(room_id=room_id, student_id=student_id)
        db.session.add(member)
        db.session.commit()
    
    messages = ChatMessage.query.filter_by(room_id=room_id).order_by(ChatMessage.created_at.asc()).limit(100).all()
    sentiment_summary = get_room_sentiment_summary(room.messages)
    
    return render_template('chat_room.html', room=room, messages=messages, sentiment_summary=sentiment_summary, student_id=student_id)

@app.route('/chat/create', methods=['GET', 'POST'])
@student_required
def create_chat_room():
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        category = request.form.get('category')
        
        room = ChatRoom(name=name, description=description, category=category, created_by=session['student_id'])
        db.session.add(room)
        db.session.commit()
        
        member = ChatRoomMember(room_id=room.id, student_id=session['student_id'])
        db.session.add(member)
        db.session.commit()
        
        log_student_action(session['student_id'], 'Create Chat Room', f'Created room: {name}')
        return redirect(url_for('chat_room', room_id=room.id))
    
    return render_template('create_chat_room.html')

# ==================== SOCKET.IO EVENTS ====================

@socketio.on('join')
def handle_join(data):
    room_id = data['room_id']
    username = data.get('username', session.get('student_name', 'Student'))
    join_room(str(room_id))
    emit('user_joined', {'username': username, 'timestamp': datetime.now().strftime('%H:%M')}, room=str(room_id))

@socketio.on('leave')
def handle_leave(data):
    room_id = data['room_id']
    username = data.get('username', session.get('student_name', 'Student'))
    leave_room(str(room_id))
    emit('user_left', {'username': username, 'timestamp': datetime.now().strftime('%H:%M')}, room=str(room_id))

@socketio.on('send_message')
def handle_send_message(data):
    room_id = data['room_id']
    student_id = session.get('student_id')

    if not student_id:
        emit('error', {'message': 'Not authenticated'})
        return

    member = ChatRoomMember.query.filter_by(room_id=room_id, student_id=student_id).first()
    if not member:
        emit('error', {'message': 'You must be a member of this room to send messages'})
        return

    message_type = data.get('message_type') or 'text'
    voice_data = None
    voice_duration = None

    # ---- Reply-to validation ----
    reply_to_id = data.get('reply_to_id')
    reply_to_msg = None
    if reply_to_id:
        reply_to_msg = ChatMessage.query.filter_by(id=reply_to_id, room_id=room_id).first()
        if not reply_to_msg:
            reply_to_id = None

    if message_type == 'voice':
        voice_data = data.get('voice_data') or ''
        voice_duration = data.get('voice_duration')
        # Guard against absurdly large payloads (~3.5MB of base64, roughly a few
        # minutes of compressed audio) so a bad client can't blow up the DB row.
        if not voice_data or len(voice_data) > 3500000:
            emit('error', {'message': 'Voice note is missing or too large to send.'})
            return
        message_text = '🎤 Voice note'
        analysis = {'cleaned_message': None, 'sentiment': 'Neutral', 'sentiment_score': 0.0,
                    'urgency_score': 1, 'is_flagged': False}
    else:
        message_text = (data.get('message') or '').strip()
        if not message_text:
            emit('error', {'message': 'Message cannot be empty'})
            return
        analysis = analyze_chat_message(message_text)

    chat_message = ChatMessage(
        room_id=room_id, student_id=student_id, message=message_text,
        cleaned_message=analysis['cleaned_message'], sentiment=analysis['sentiment'],
        sentiment_score=analysis['sentiment_score'], urgency_score=analysis['urgency_score'],
        is_flagged=analysis['is_flagged'], reply_to_id=reply_to_id,
        message_type=message_type, voice_data=voice_data, voice_duration=voice_duration
    )
    db.session.add(chat_message)
    
    room = ChatRoom.query.get(room_id)
    if room:
        room.last_activity = datetime.utcnow()
    
    db.session.commit()
    
    student = Student.query.get(student_id)
    student_name = student.full_name.split()[0] if student else "Student"

    reply_to_payload = None
    if reply_to_msg:
        rt_student = Student.query.get(reply_to_msg.student_id)
        rt_name = rt_student.full_name.split()[0] if rt_student else 'Student'
        reply_to_payload = {
            'id': reply_to_msg.id,
            'username': rt_name,
            'preview': '🎤 Voice note' if reply_to_msg.message_type == 'voice' else (reply_to_msg.message or '')[:80]
        }

    payload = {
        'id': chat_message.id,
        'room_id': room_id,
        'room_name': room.name if room else None,
        'message': message_text,
        'message_type': message_type,
        'voice_data': voice_data,
        'voice_duration': voice_duration,
        'reply_to': reply_to_payload,
        'username': student_name,
        'sentiment': analysis['sentiment'],
        'sentiment_score': analysis['sentiment_score'],
        'urgency_score': analysis['urgency_score'],
        'is_flagged': analysis['is_flagged'],
        'timestamp': chat_message.created_at.strftime('%Y-%m-%d %H:%M')
    }

    # Broadcast the new message to the room participants.
    emit('new_message', payload, room=str(room_id))

    # Also stream it live to the admin chat messages page.
    emit('admin_new_message', payload, room='admins')


    # Update daily sentiment stats
    today = datetime.utcnow().date()
    start_of_day = datetime(today.year, today.month, today.day)
    end_of_day = start_of_day + timedelta(days=1)
    
    messages_today = ChatMessage.query.filter(
        ChatMessage.room_id == room_id,
        ChatMessage.created_at >= start_of_day,
        ChatMessage.created_at < end_of_day
    ).all()
    
    positive = sum(1 for m in messages_today if m.sentiment == 'Positive')
    negative = sum(1 for m in messages_today if m.sentiment == 'Negative')
    neutral = len(messages_today) - positive - negative
    avg_urgency = sum(m.urgency_score for m in messages_today) / len(messages_today) if messages_today else 0
    
    stats = ChatRoomSentiment.query.filter_by(room_id=room_id, date=today).first()
    if stats:
        stats.positive_count = positive
        stats.negative_count = negative
        stats.neutral_count = neutral
        stats.avg_urgency = avg_urgency
        stats.total_messages = len(messages_today)
    else:
        stats = ChatRoomSentiment(
            room_id=room_id, date=today, positive_count=positive,
            negative_count=negative, neutral_count=neutral,
            avg_urgency=avg_urgency, total_messages=len(messages_today)
        )
        db.session.add(stats)
    db.session.commit()

# ==================== ANNOUNCEMENTS ====================


@app.route('/announcements')
def announcements():
    now = datetime.utcnow()
    active_announcements = Announcement.query.filter(
        Announcement.is_active == True,
        (Announcement.expires_at.is_(None) | (Announcement.expires_at > now))
    ).order_by(Announcement.created_at.desc()).all()
    return render_template('announcements.html', announcements=active_announcements)

@app.route('/admin/announcements', methods=['GET', 'POST'])
@src_required
def admin_announcements():
    if request.method == 'POST':
        title = request.form.get('title')
        content = request.form.get('content')
        expires_days = request.form.get('expires_days', type=int)
        
        announcement = Announcement(
            title=title, content=content, author=session['admin_name'],
            expires_at=datetime.utcnow() + timedelta(days=expires_days) if expires_days else None
        )
        db.session.add(announcement)
        db.session.commit()
        
        log_admin_action(session['admin_name'], 'Create Announcement', f'Title: {title}')
        return redirect(url_for('admin_announcements'))
    
    announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
    return render_template('admin_announcements.html', announcements=announcements, datetime=datetime)

@app.route('/admin/announcements/delete/<int:id>', methods=['POST'])
@src_required
def delete_announcement(id):
    announcement = Announcement.query.get_or_404(id)
    db.session.delete(announcement)
    db.session.commit()
    return redirect(url_for('admin_announcements'))

# ==================== SRC ADMIN ROUTES ====================

@app.route('/admin/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        totp_token = request.form.get('totp_token', '')
        
        user = SRCUser.query.filter_by(username=username).first()
        ip_address = get_client_ip()
        user_agent = get_user_agent()
        
        if user and check_password_hash(user.password_hash, password):
            # Check for anomalies
            anomaly = SecurityManager.detect_anomaly(user.id, ip_address, user_agent)
            
            # If 2FA is enabled, verify token
            if user.is_2fa_enabled:
                if not totp_token:
                    session['pending_2fa_user_id'] = user.id
                    return render_template('admin_2fa.html', username=username)
                if not SecurityManager.verify_2fa_token(user.totp_secret, totp_token):
                    SecurityManager.record_login(user.id, 'admin', ip_address, user_agent, False, 'Invalid 2FA token')
                    return render_template('admin_2fa.html', username=username, error='Invalid 2FA code')
            
            # Login successful
            session['admin_id'] = user.id
            session['admin_name'] = user.full_name
            session['admin_role'] = user.role
            session['2fa_verified'] = True
            session.pop('pending_2fa_user_id', None)
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            SecurityManager.record_login(user.id, 'admin', ip_address, user_agent, True)
            log_admin_action(user.full_name, 'Login Success', f'Logged in from {ip_address}')
            
            if anomaly['is_anomaly']:
                flash(anomaly['message'], 'warning')
            
            return redirect(url_for('admin_dashboard'))
        else:
            if user:
                SecurityManager.record_login(user.id, 'admin', ip_address, user_agent, False, 'Invalid password')
            return render_template('admin_login.html', error='Invalid credentials')
    
    return render_template('admin_login.html')

@app.route('/admin/2fa/setup', methods=['GET', 'POST'])
@src_required
def admin_2fa_setup():
    """Setup 2FA for admin account."""
    user = SRCUser.query.get(session['admin_id'])
    if request.method == 'POST':
        token = request.form.get('token', '')
        secret = session.get('2fa_setup_secret')
        if secret and SecurityManager.verify_2fa_token(secret, token):
            user.totp_secret = secret
            user.is_2fa_enabled = True
            db.session.commit()
            session.pop('2fa_setup_secret', None)
            flash('Two-factor authentication enabled successfully', 'success')
            return redirect(url_for('admin_dashboard'))
        return render_template('admin_2fa_setup.html', error='Invalid verification code', secret=secret)
    
    secret = SecurityManager.generate_2fa_secret()
    session['2fa_setup_secret'] = secret
    totp_uri = SecurityManager.get_totp_uri(secret, user.username)
    return render_template('admin_2fa_setup.html', secret=secret, totp_uri=totp_uri)

@app.route('/admin/2fa/verify', methods=['GET', 'POST'])
def admin_2fa_verify():
    """Verify 2FA token during login."""
    if request.method == 'POST':
        username = request.form.get('username')
        totp_token = request.form.get('totp_token', '')
        user = SRCUser.query.filter_by(username=username).first()
        ip_address = get_client_ip()
        user_agent = get_user_agent()
        
        if user and user.is_2fa_enabled and SecurityManager.verify_2fa_token(user.totp_secret, totp_token):
            session['admin_id'] = user.id
            session['admin_name'] = user.full_name
            session['admin_role'] = user.role
            session['2fa_verified'] = True
            session.pop('pending_2fa_user_id', None)
            user.last_login = datetime.utcnow()
            db.session.commit()
            SecurityManager.record_login(user.id, 'admin', ip_address, user_agent, True)
            return redirect(url_for('admin_dashboard'))
        
        SecurityManager.record_login(user.id if user else 0, 'admin', ip_address, user_agent, False, 'Invalid 2FA')
        return render_template('admin_2fa.html', username=username, error='Invalid 2FA code')
    
    return render_template('admin_2fa.html')

@app.route('/admin/2fa/disable', methods=['POST'])
@src_required
def admin_2fa_disable():
    """Disable 2FA for admin account."""
    user = SRCUser.query.get(session['admin_id'])
    user.totp_secret = None
    user.is_2fa_enabled = False
    db.session.commit()
    flash('Two-factor authentication disabled', 'info')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/dashboard')
@src_required
def admin_dashboard():
    page = request.args.get('page', 1, type=int)

    per_page = request.args.get('per_page', 20, type=int)
    category_filter = request.args.get('category', '')
    status_filter = request.args.get('status', '')
    urgency_filter = request.args.get('urgency', '')
    search_query = request.args.get('search', '')
    
    query = Feedback.query
    if category_filter:
        query = query.filter_by(category=category_filter)
    if status_filter:
        query = query.filter_by(status=status_filter)
    if urgency_filter:
        query = query.filter(Feedback.urgency_score >= int(urgency_filter))
    if search_query:
        query = query.filter(Feedback.feedback_text.contains(search_query))
    
    paginated = query.order_by(Feedback.urgency_score.desc(), Feedback.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    total = Feedback.query.count()
    urgent = Feedback.query.filter(Feedback.urgency_score >= 4).count()
    resolved = Feedback.query.filter_by(status='Resolved').count()
    positive = Feedback.query.filter_by(sentiment='Positive').count()
    negative = Feedback.query.filter_by(sentiment='Negative').count()
    neutral = Feedback.query.filter_by(sentiment='Neutral').count()
    total_students = Student.query.count()
    profanity_count = Feedback.query.filter_by(has_profanity=True).count()

    
    trend_data = []
    for i in range(6):
        month = datetime.utcnow().replace(day=1) - timedelta(days=30*i)
        month_feedbacks = Feedback.query.filter(
            Feedback.created_at >= month,
            Feedback.created_at < month.replace(day=28) + timedelta(days=4)
        ).all()
        trend_data.append({
            'month': month.strftime('%b'),
            'positive': sum(1 for f in month_feedbacks if f.sentiment == 'Positive'),
            'negative': sum(1 for f in month_feedbacks if f.sentiment == 'Negative'),
            'neutral': sum(1 for f in month_feedbacks if f.sentiment == 'Neutral')
        })
    
    categories = {}
    for f in Feedback.query.all():
        categories[f.category] = categories.get(f.category, 0) + 1
    
    all_feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    duplicate_ids = _duplicate_feedback_ids(all_feedbacks[:200])
    low_confidence = sum(1 for f in all_feedbacks if (f.confidence_score or 0) < 55)
    open_negative = sum(1 for f in all_feedbacks
                        if f.sentiment == 'Negative' and f.status not in ('Resolved',))
    topic_summary = _dashboard_topic_summary(all_feedbacks)

    stats = {
        'total': total, 'urgent': urgent, 'resolved': resolved,
        'positive': positive, 'negative': negative, 'neutral': neutral,
        'resolution_rate': round((resolved / total * 100) if total > 0 else 0, 1),
        'total_students': total_students, 'profanity_count': profanity_count,
        'low_confidence': low_confidence, 'open_negative': open_negative,
        'duplicate_reports': len(duplicate_ids)
    }
    
    # SLA tracking - count issues approaching or past deadline
    now = datetime.utcnow()
    sla_warning = Feedback.query.filter(
        Feedback.status.notin_(['Resolved', 'Closed']),
        Feedback.created_at <= now - timedelta(hours=48),
        Feedback.created_at > now - timedelta(hours=72)
    ).count()
    sla_breached = Feedback.query.filter(
        Feedback.status.notin_(['Resolved', 'Closed']),
        Feedback.created_at <= now - timedelta(hours=72)
    ).count()
    stats['sla_warning'] = sla_warning
    stats['sla_breached'] = sla_breached
    
    # Precompute explanations for Analyze modal.
    # We use cleaned_text when available, fallback to feedback_text.
    explanations_by_id = {}
    recommendations_by_id = {}

    for f in paginated.items:
        base_text = (f.cleaned_text or '').strip() or (f.feedback_text or '').strip()
        sentiment_expl = get_sentiment_explanation(base_text).get('sentiment_explanation', [])
        urgency_expl = get_urgency_explanation(base_text, f.sentiment)

        explanations_by_id[f.id] = {
            'sentiment_explanation': sentiment_expl,
            'urgency_explanation': urgency_expl,
            'confidence_score': f.confidence_score,
            'dominant_emotion': f.dominant_emotion,
            'compound_mood': f.compound_mood,
            'emotion_emoji': get_emotion_emoji(f.dominant_emotion, f.compound_mood),
            'emotion_intensities': _parse_json_field(f.emotion_intensities, {}),
            'secondary_emotions': _parse_json_field(f.secondary_emotions, []),
        }

        helpful_count = SolutionFeedback.query.filter_by(feedback_id=f.id, was_helpful=True).count()
        unhelpful_count = SolutionFeedback.query.filter_by(feedback_id=f.id, was_helpful=False).count()

        recommendations_by_id[f.id] = {
            'recommended_keywords': f.recommended_keywords,
            'short_term_solution': f.short_term_solution,
            'long_term_solution': f.long_term_solution,
            'responsible_department': f.responsible_department,
            'estimated_time': f.estimated_time,
            'confidence': f.recommendation_confidence,
            'secondary_categories': _parse_json_field(f.secondary_categories, []),
            'used_template_id': f.used_template_id,
            'helpful_count': helpful_count,
            'unhelpful_count': unhelpful_count,
        }

    # Enhanced recommendation data
    trending_issues = get_trending_issues()
    department_workload = get_department_workload()

    return render_template('admin_dashboard.html', feedbacks=paginated.items, pagination=paginated,
                         stats=stats, categories=categories, trend_data=trend_data,
                         admin_name=session.get('admin_name'), current_category=category_filter,
                         current_status=status_filter, current_urgency=urgency_filter, current_search=search_query,
                         explanations_by_id=explanations_by_id,
                         recommendations_by_id=recommendations_by_id,
                         topic_summary=topic_summary, duplicate_ids=duplicate_ids,
                         trending_issues=trending_issues, department_workload=department_workload)



@app.route('/admin/update/<int:feedback_id>', methods=['POST'])
@src_required
def update_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    old_status = feedback.status
    new_status = request.form.get('status', feedback.status)
    
    feedback.status = new_status
    feedback.assigned_to = request.form.get('assigned_to', feedback.assigned_to)
    feedback.src_response = request.form.get('src_response', feedback.src_response)
    
    if feedback.status == 'Resolved' and not feedback.resolved_at:
        feedback.resolved_at = datetime.utcnow()
    
    db.session.commit()
    
    log_admin_action(session['admin_name'], 'Feedback Update', f'Feedback ID: {feedback_id}, Status: {old_status} -> {new_status}')
    add_db_log('feedback', 'INFO', 'admin', session['admin_name'], 'Feedback Updated', f'Feedback ID: {feedback_id}, Status: {old_status} -> {new_status}')
    
    # Create notification for student
    if feedback.student_id:
        if old_status != new_status:
            status_messages = {
                'Pending': 'Your feedback is pending review.',
                'Acknowledged': 'Your feedback has been acknowledged by the SRC.',
                'In Progress': 'The SRC is working on your feedback.',
                'Resolved': 'Your feedback has been resolved!',
                'Escalated': 'Your feedback has been escalated to the relevant department.'
            }
            create_notification(
                student_id=feedback.student_id,
                feedback_id=feedback.id,
                notification_type='status_change',
                title=f'Status Updated: {new_status}',
                message=status_messages.get(new_status, f'Your feedback status was updated to {new_status}.')
            )
        if request.form.get('src_response') and request.form.get('src_response') != feedback.src_response:
            create_notification(
                student_id=feedback.student_id,
                feedback_id=feedback.id,
                notification_type='response',
                title='New SRC Response',
                message='The SRC has responded to your feedback.'
            )
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/feedback/<int:feedback_id>/regenerate-recommendation', methods=['POST'])
@src_required
def regenerate_recommendation(feedback_id):
    """Recompute the solution recommendation for a feedback item on demand.

    Useful when an admin reclassifies the category, or when new
    SolutionTemplate overrides have been added since the item was
    originally submitted -- the stored recommendation would otherwise
    stay frozen at whatever it was on submission day.
    """
    feedback = Feedback.query.get_or_404(feedback_id)
    base_text = (feedback.cleaned_text or '').strip() or (feedback.feedback_text or '').strip()
    emotion_data = {
        'dominant_emotion': feedback.dominant_emotion,
        'compound_mood': feedback.compound_mood,
    }

    rec_result = generate_recommendation(
        text=base_text,
        category=feedback.category,
        urgency_score=feedback.urgency_score,
        sentiment=feedback.sentiment,
        sentiment_score=feedback.sentiment_score,
        emotion=emotion_data,
    )
    _apply_enhanced_recommendation(feedback, rec_result)
    db.session.commit()

    log_admin_action(session['admin_name'], 'Recommendation Regenerated', f'Feedback ID: {feedback_id}')

    return jsonify({
        'success': True,
        'recommended_keywords': feedback.recommended_keywords,
        'short_term_solution': feedback.short_term_solution,
        'long_term_solution': feedback.long_term_solution,
        'responsible_department': feedback.responsible_department,
        'estimated_time': feedback.estimated_time,
        'confidence': feedback.recommendation_confidence,
        'secondary_categories': _parse_json_field(feedback.secondary_categories, []),
    })

@app.route('/admin/feedback/bulk-action', methods=['POST'])
@src_required
def bulk_action():
    """Perform bulk actions on multiple feedback items."""
    data = request.get_json()
    action = data.get('action', '')
    ids = data.get('ids', [])
    
    if not ids:
        return jsonify({'success': False, 'error': 'No items selected'})
    
    feedbacks = Feedback.query.filter(Feedback.id.in_(ids)).all()
    count = 0
    
    for f in feedbacks:
        if action == 'resolve':
            f.status = 'Resolved'
            if not f.resolved_at:
                f.resolved_at = datetime.utcnow()
        elif action == 'pending':
            f.status = 'Pending'
        elif action == 'in_progress':
            f.status = 'In Progress'
        elif action == 'delete':
            db.session.delete(f)
        elif action.startswith('assign_'):
            f.assigned_to = action[7:]
        elif action.startswith('category_'):
            f.category = action[9:]
        count += 1
    
    db.session.commit()
    log_admin_action(session['admin_name'], 'Bulk Action', f'{action} applied to {count} items')
    
    return jsonify({'success': True, 'count': count})

@app.route('/admin/feedback/<int:feedback_id>/solution-feedback', methods=['POST'])
@src_required
def solution_feedback(feedback_id):
    """Record whether a solution was helpful."""
    data = request.get_json()
    was_helpful = data.get('was_helpful', False)
    
    feedback = Feedback.query.get_or_404(feedback_id)
    
    # Record the feedback
    sf = SolutionFeedback(
        feedback_id=feedback_id,
        template_category=feedback.category,
        was_helpful=was_helpful,
        created_by=session.get('admin_name')
    )
    db.session.add(sf)
    db.session.commit()
    
    # Update recommendation engine
    solution_key = (feedback.short_term_solution or '')[:50]
    get_engine().record_solution_feedback(feedback.category, solution_key, was_helpful, feedback.status == 'Resolved')
    
    helpful = SolutionFeedback.query.filter_by(feedback_id=feedback_id, was_helpful=True).count()
    unhelpful = SolutionFeedback.query.filter_by(feedback_id=feedback_id, was_helpful=False).count()
    
    return jsonify({'success': True, 'helpful_count': helpful, 'unhelpful_count': unhelpful})

@app.route('/admin/feedback/<int:feedback_id>/solution-feedback', methods=['POST'])
@src_required
def rate_solution_feedback(feedback_id):
    """Admin marks a recommendation as helpful/unhelpful.

    Feeds the SolutionFeedback table, which RecommendationLearner uses to
    score keyword and template effectiveness over time -- this is the
    data source for the Recommendation Insights panel in Analytics.
    """
    feedback = Feedback.query.get_or_404(feedback_id)
    data = request.get_json(silent=True) or request.form
    was_helpful = str(data.get('was_helpful')).lower() in ('true', '1', 'yes')
    comment = (data.get('comment') or '').strip() or None

    resolution_hours = None
    if feedback.status == 'Resolved' and feedback.resolved_at:
        resolution_hours = round((feedback.resolved_at - feedback.created_at).total_seconds() / 3600, 1)

    entry = SolutionFeedback(
        feedback_id=feedback.id,
        template_category=feedback.category,
        was_helpful=was_helpful,
        resolved_after=(feedback.status == 'Resolved'),
        resolution_time_hours=resolution_hours,
        comment=comment,
        created_by=session.get('admin_name'),
    )
    db.session.add(entry)

    if feedback.used_template_id and was_helpful:
        tmpl = SolutionTemplate.query.get(feedback.used_template_id)
        if tmpl:
            tmpl.resolution_count = (tmpl.resolution_count or 0) + 1

    db.session.commit()

    helpful_count = SolutionFeedback.query.filter_by(feedback_id=feedback.id, was_helpful=True).count()
    unhelpful_count = SolutionFeedback.query.filter_by(feedback_id=feedback.id, was_helpful=False).count()

    return jsonify({'success': True, 'helpful_count': helpful_count, 'unhelpful_count': unhelpful_count})

@app.route('/admin/solution-templates', methods=['GET', 'POST'])
@src_required
def admin_solution_templates():
    """Admin-facing CRUD for SolutionTemplate overrides.

    Any active template here takes precedence over the built-in
    templates in solution_recommender.py for its category, so admins
    can tailor recommendations to their institution without a code
    change.
    """
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'create':
            category = request.form.get('category', '').strip()
            keywords = request.form.get('keywords', '').strip()
            short_term = request.form.get('short_term_solution', '').strip()
            long_term = request.form.get('long_term_solution', '').strip()
            department = request.form.get('responsible_department', '').strip()
            estimated_time = request.form.get('estimated_time', '').strip()

            if category and keywords and short_term and long_term and department and estimated_time:
                tmpl = SolutionTemplate(
                    category=category, keywords=keywords,
                    short_term_solution=short_term, long_term_solution=long_term,
                    responsible_department=department, estimated_time=estimated_time,
                    created_by=session.get('admin_name'),
                )
                db.session.add(tmpl)
                db.session.commit()
                log_admin_action(session['admin_name'], 'Solution Template Created', f'Category: {category}')

        elif action == 'toggle':
            tmpl_id = request.form.get('template_id', type=int)
            tmpl = SolutionTemplate.query.get(tmpl_id) if tmpl_id else None
            if tmpl:
                tmpl.is_active = not tmpl.is_active
                db.session.commit()
                log_admin_action(session['admin_name'], 'Solution Template Toggled',
                                  f'ID: {tmpl_id}, Active: {tmpl.is_active}')

        elif action == 'delete':
            tmpl_id = request.form.get('template_id', type=int)
            tmpl = SolutionTemplate.query.get(tmpl_id) if tmpl_id else None
            if tmpl:
                db.session.delete(tmpl)
                db.session.commit()
                log_admin_action(session['admin_name'], 'Solution Template Deleted', f'ID: {tmpl_id}')

        return redirect(url_for('admin_solution_templates'))

    templates = SolutionTemplate.query.order_by(SolutionTemplate.category, SolutionTemplate.created_at.desc()).all()
    categories = ['Accommodation', 'ICT/Wi-Fi', 'Academics', 'Catering', 'Facilities', 'Safety',
                  'Transport', 'Mental Health', 'Financial', 'Administration', 'Cleanliness', 'Other']

    return render_template('admin_solution_templates.html', templates=templates, categories=categories)

@app.route('/admin/templates', methods=['GET', 'POST'])
@src_required
def admin_templates():
    if request.method == 'POST':
        template_key = request.form.get('template_key')
        template_value = request.form.get('template_value')
        if template_key and template_value:
            RESPONSE_TEMPLATES[template_key] = template_value
            log_admin_action(session['admin_name'], 'Template Update', f'Updated: {template_key}')
    
    return render_template('admin_templates.html', templates=RESPONSE_TEMPLATES)

@app.route('/admin/delete-feedback/<int:feedback_id>', methods=['POST'])
@src_required
def admin_delete_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    db.session.delete(feedback)
    db.session.commit()
    log_admin_action(session['admin_name'], 'Delete Feedback', f'Deleted feedback ID: {feedback_id}')
    add_db_log('feedback', 'INFO', 'admin', session['admin_name'], 'Feedback Deleted', f'Feedback ID: {feedback_id}')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/students')
@src_required
def admin_students():
    students = Student.query.order_by(Student.created_at.desc()).all()
    return render_template('admin_students.html', students=students)


@app.route('/admin/ai-review')
@src_required
def admin_ai_review():
    """Phase 2: queue feedback that needs administrator review."""
    threshold = request.args.get('threshold', 70, type=float)
    query = Feedback.query.filter(
        or_(
            Feedback.confidence_score < threshold,
            Feedback.confidence_score.is_(None),
        )
    ).order_by(Feedback.urgency_score.desc(), Feedback.created_at.desc())
    reviews = query.limit(200).all()
    for f in reviews:
        base_text = (f.cleaned_text or '').strip() or (f.feedback_text or '').strip()
        f.phase2_ai_explanation = build_ai_explanation(
            base_text,
            category=f.category,
            recommendation={'short_term_solution': f.short_term_solution},
            final_sentiment=f.sentiment,
            final_confidence=f.confidence_score,
        )
    return render_template(
        'admin_ai_review.html',
        reviews=reviews,
        threshold=threshold,
        review_count=query.count(),
        categories=sorted({c for c in [f.category for f in reviews] if c}),
    )


@app.route('/admin/ai-review/<int:feedback_id>', methods=['POST'])
@src_required
def review_ai_feedback(feedback_id):
    """Approve/correct an AI result and record the action."""
    feedback = Feedback.query.get_or_404(feedback_id)
    old = (feedback.sentiment, feedback.category, feedback.urgency_score)
    new_sentiment = (request.form.get('sentiment') or feedback.sentiment or 'neutral').strip().lower()
    new_category = (request.form.get('category') or feedback.category or 'general').strip()
    new_urgency = max(1, min(5, request.form.get('urgency', feedback.urgency_score, type=int)))
    notes = (request.form.get('notes') or '').strip()[:2000]

    allowed = {'positive', 'negative', 'neutral'}
    if new_sentiment not in allowed:
        new_sentiment = feedback.sentiment or 'neutral'

        feedback.sentiment = new_sentiment
        feedback.category = new_category
        feedback.urgency_score = new_urgency

        # Regenerate solution recommendation if sentiment changed to Negative
        if new_sentiment == 'negative' and old[0] != 'negative':
            try:
                rec_result = generate_recommendation(
                    text=feedback.feedback_text,
                    category=new_category,
                    urgency_score=new_urgency,
                    sentiment=new_sentiment,
                    sentiment_score=-0.5,
                )
                _apply_enhanced_recommendation(feedback, rec_result)
                feedback.secondary_categories = json.dumps(rec.secondary_categories) if rec.secondary_categories else '[]'
                feedback.recommendation_confidence = rec.confidence
            except Exception:
                pass

        action = 'approved' if old == (new_sentiment, new_category, new_urgency) else 'corrected'
    db.session.add(AIReviewLog(
        feedback_id=feedback.id, admin_name=session.get('admin_name', 'admin'),
        action=action, old_sentiment=old[0], new_sentiment=new_sentiment,
        old_category=old[1], new_category=new_category,
        old_urgency=old[2], new_urgency=new_urgency, notes=notes
    ))
    # Active-learning loop: teach the custom lexicon any domain words the
    # generic engines missed, scored by the human reviewer's label. Words are
    # added to the session and persisted by the commit below.
    if new_sentiment in ('positive', 'negative'):
        try:
            from sentiment.custom_lexicon import CustomLexiconManager
            learned = CustomLexiconManager().learn_from_correction(
                (feedback.feedback_text or ''), new_sentiment,
                session.get('admin_name', 'admin'),
            )
            if learned:
                log_admin_action(
                    session['admin_name'], 'AI Review',
                    f'Feedback ID {feedback.id}: auto-learned {learned} lexicon term(s)',
                )
        except Exception:
            pass
    
    # Record correction in SentimentCorrection table for active learning stats
    if old[0] != new_sentiment:
        try:
            from database import SentimentCorrection
            correction = SentimentCorrection(
                feedback_id=feedback.id,
                original_sentiment=old[0],
                corrected_sentiment=new_sentiment,
                admin_name=session.get('admin_name', 'admin'),
                confidence_before=feedback.confidence_score or 0.0,
            )
            db.session.add(correction)
        except Exception:
            pass
    
    db.session.commit()
    log_admin_action(session['admin_name'], 'AI Review', f'Feedback ID {feedback.id}: {action}')

    # Create notification for student
    if feedback.student_id:
        if old[0] != new_sentiment:
            create_notification(
                student_id=feedback.student_id,
                feedback_id=feedback.id,
                notification_type='sentiment_change',
                title='Feedback Sentiment Updated',
                message=f'Your feedback #{feedback.id} sentiment was changed from {old[0]} to {new_sentiment}.'
            )
        if old[1] != new_category:
            create_notification(
                student_id=feedback.student_id,
                feedback_id=feedback.id,
                notification_type='category_change',
                title='Feedback Category Updated',
                message=f'Your feedback #{feedback.id} category was changed from {old[1]} to {new_category}.'
            )

    return redirect(url_for('admin_ai_review'))


@app.route('/admin/ai-review/<int:feedback_id>/add-to-lexicon', methods=['POST'])
@src_required
def review_add_to_lexicon(feedback_id):
    """Add an admin-approved term from a reviewed feedback item."""
    feedback = Feedback.query.get_or_404(feedback_id)
    word = (request.form.get('word') or '').strip().lower()
    score = request.form.get('score', type=float)
    category = (request.form.get('lexicon_category') or 'general').strip()[:50]
    if not word or score is None or not -1 <= score <= 1:
        return redirect(url_for('admin_ai_review'))
    row = CustomLexicon.query.filter_by(word=word).first()
    if row:
        row.sentiment_score, row.category, row.is_active = score, category, True
        row.added_by = session.get('admin_name')
    else:
        db.session.add(CustomLexicon(word=word, sentiment_score=score,
                                     category=category, added_by=session.get('admin_name')))
    db.session.add(AIReviewLog(
        feedback_id=feedback.id, admin_name=session.get('admin_name', 'admin'),
        action='lexicon_add', notes=f'Added "{word}" with score {score:.2f}'
    ))
    db.session.commit()
    return redirect(url_for('admin_ai_review'))


@app.route('/admin/lexicon-manager', methods=['GET', 'POST'])
@src_required
def admin_lexicon_manager():
    """Phase 2: manage the database-backed custom lexicon without editing code."""
    if request.method == 'POST':
        word = (request.form.get('word') or '').strip().lower()
        score = request.form.get('sentiment_score', type=float)
        category = (request.form.get('category') or 'general').strip()[:50]
        if not word or score is None or not -1 <= score <= 1:
            return redirect(url_for('admin_lexicon_manager', error='invalid'))
        row = CustomLexicon.query.filter_by(word=word).first()
        if row:
            row.sentiment_score, row.category, row.is_active = score, category, True
            row.added_by = session.get('admin_name')
        else:
            row = CustomLexicon(word=word, sentiment_score=score, category=category,
                                added_by=session.get('admin_name'))
            db.session.add(row)
        db.session.commit()
        log_admin_action(session['admin_name'], 'Lexicon Update', f'Word "{word}" score {score}')
        return redirect(url_for('admin_lexicon_manager'))

    entries = CustomLexicon.query.order_by(CustomLexicon.updated_at.desc()).all()
    unknowns = UnknownWord.query.filter_by(is_reviewed=False).order_by(UnknownWord.created_at.desc()).limit(50).all()
    return render_template('admin_lexicon_manager.html', entries=entries, unknowns=unknowns,
                           error=request.args.get('error'))


@app.route('/admin/lexicon-manager/<int:entry_id>/toggle', methods=['POST'])
@src_required
def admin_lexicon_toggle(entry_id):
    entry = CustomLexicon.query.get_or_404(entry_id)
    entry.is_active = not entry.is_active
    db.session.commit()
    log_admin_action(session['admin_name'], 'Lexicon Toggle',
                     f'Word "{entry.word}" active={entry.is_active}')
    return redirect(url_for('admin_lexicon_manager'))


@app.route('/admin/lexicon-manager/<int:entry_id>/delete', methods=['POST'])
@src_required
def admin_lexicon_delete(entry_id):
    entry = CustomLexicon.query.get_or_404(entry_id)
    word = entry.word
    db.session.delete(entry)
    db.session.commit()
    log_admin_action(session['admin_name'], 'Lexicon Delete', f'Word "{word}"')
    return redirect(url_for('admin_lexicon_manager'))


@app.route('/admin/ai-audit')
@src_required
def admin_ai_audit():
    logs = AIReviewLog.query.order_by(AIReviewLog.created_at.desc()).limit(200).all()
    return render_template('admin_ai_audit.html', logs=logs)


@app.route('/admin/lexicon-gaps')
@src_required
def admin_lexicon_gaps():
    from sentiment.unknown_detector import UnknownWordDetector

    # Cap the scan to the most recent 1000 items
    recent_feedback = (
        Feedback.query
        .order_by(Feedback.id.desc())
        .limit(1000)
        .all()
    )
    feedback_items = [(f.feedback_text, f.sentiment) for f in recent_feedback]

    detector = UnknownWordDetector()
    report = detector.scan_feedback_for_gaps(feedback_items, top_n=40)

    # Add sentiment suggestions to gap words
    def suggest_score(word, example_text):
        """Simple sentiment suggestion based on context words."""
        negative_words = ['bad', 'terrible', 'worst', 'hate', 'awful', 'poor', 'broken', 'not', 'no', 'never', 'slow', 'problem', 'issue', 'fail', 'wrong', 'difficult', 'hard', 'lack', 'need', 'missing']
        positive_words = ['good', 'great', 'best', 'love', 'excellent', 'amazing', 'wonderful', 'perfect', 'nice', 'happy', 'easy', 'fast', 'improve', 'better', 'awesome', 'helpful']
        
        text_lower = example_text.lower()
        neg_count = sum(1 for w in negative_words if w in text_lower)
        pos_count = sum(1 for w in positive_words if w in text_lower)
        
        if neg_count > pos_count:
            return round(-0.3 * neg_count, 1)
        elif pos_count > neg_count:
            return round(0.3 * pos_count, 1)
        return 0.0

    # Enrich gap words with suggestions
    for item in report.get('neutral_gap_words', []):
        item['suggested_score'] = suggest_score(item['word'], item.get('example', ''))
    
    for item in report.get('all_gap_words', []):
        item['suggested_score'] = suggest_score(item['word'], item.get('example', ''))

    return render_template('admin_lexicon_gaps.html',
                           scanned_count=report['scanned_count'],
                           total_feedback=len(recent_feedback),
                           neutral_gap_words=report['neutral_gap_words'],
                           all_gap_words=report['all_gap_words'])

@app.route('/admin/api/lexicon/add-word', methods=['POST'])
@src_required
def api_lexicon_add_word():
    """Add a single word to the custom lexicon."""
    data = request.get_json()
    word = data.get('word', '').strip().lower()
    score = float(data.get('sentiment_score', 0))
    category = data.get('category', 'general')
    
    if not word:
        return jsonify({'success': False, 'error': 'Word is required'})
    
    # Check if word already exists
    existing = CustomLexicon.query.filter_by(word=word).first()
    if existing:
        existing.sentiment_score = score
        existing.category = category
        existing.is_active = True
    else:
        entry = CustomLexicon(
            word=word,
            sentiment_score=score,
            category=category,
            is_active=True
        )
        db.session.add(entry)
    
    db.session.commit()
    log_admin_action(session['admin_name'], 'Lexicon Add', f'Word "{word}" score={score}')
    return jsonify({'success': True})

@app.route('/admin/api/lexicon/add-words-bulk', methods=['POST'])
@src_required
def api_lexicon_add_words_bulk():
    """Add multiple words to the custom lexicon."""
    data = request.get_json()
    words = data.get('words', [])
    
    count = 0
    for item in words:
        word = item.get('word', '').strip().lower()
        score = float(item.get('sentiment_score', 0))
        if not word:
            continue
        
        existing = CustomLexicon.query.filter_by(word=word).first()
        if existing:
            existing.sentiment_score = score
            existing.is_active = True
        else:
            entry = CustomLexicon(word=word, sentiment_score=score, category='general', is_active=True)
            db.session.add(entry)
        count += 1
    
    db.session.commit()
    log_admin_action(session['admin_name'], 'Lexicon Bulk Add', f'{count} words added')
    return jsonify({'success': True, 'count': count})

@app.route('/admin/api/ai-review/<int:feedback_id>/quick-approve', methods=['POST'])
@src_required
def api_ai_review_quick_approve(feedback_id):
    """Quick approve an AI classification."""
    feedback = Feedback.query.get_or_404(feedback_id)
    # Mark as reviewed by creating an audit log
    log = AIReviewLog(
        feedback_id=feedback_id,
        action='approved',
        admin_name=session.get('admin_name'),
        notes='Quick approved via review queue'
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/api/ai-review/<int:feedback_id>/quick-reject', methods=['POST'])
@src_required
def api_ai_review_quick_reject(feedback_id):
    """Quick reject an AI classification."""
    feedback = Feedback.query.get_or_404(feedback_id)
    log = AIReviewLog(
        feedback_id=feedback_id,
        action='rejected',
        admin_name=session.get('admin_name'),
        notes='Quick rejected via review queue'
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/api/ai-review/bulk-approve', methods=['POST'])
@src_required
def api_ai_review_bulk_approve():
    """Bulk approve multiple AI classifications."""
    data = request.get_json()
    ids = data.get('ids', [])
    
    count = 0
    for fid in ids:
        feedback = Feedback.query.get(fid)
        if feedback:
            log = AIReviewLog(
                feedback_id=fid,
                action='approved',
                admin_name=session.get('admin_name'),
                notes='Bulk approved via review queue'
            )
            db.session.add(log)
            count += 1
    
    db.session.commit()
    log_admin_action(session['admin_name'], 'AI Review Bulk Approve', f'{count} items approved')
    return jsonify({'success': True, 'count': count})

# ==================== NOTIFICATION API ROUTES ====================

@app.route('/api/notifications')
@login_required
def api_get_notifications():
    """Get notifications for the logged-in student."""
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    
    notifications = Notification.query.filter_by(student_id=student_id).order_by(Notification.created_at.desc()).limit(20).all()
    unread_count = Notification.query.filter_by(student_id=student_id, is_read=False).count()
    
    return jsonify({
        'success': True,
        'notifications': [n.to_dict() for n in notifications],
        'unread_count': unread_count
    })

@app.route('/api/notifications/mark-read', methods=['POST'])
@login_required
def api_mark_notifications_read():
    """Mark notifications as read."""
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    
    data = request.get_json()
    notification_ids = data.get('ids', [])
    
    if notification_ids:
        Notification.query.filter(
            Notification.id.in_(notification_ids),
            Notification.student_id == student_id
        ).update({'is_read': True}, synchronize_session=False)
    else:
        # Mark all as read
        Notification.query.filter_by(student_id=student_id, is_read=False).update({'is_read': True}, synchronize_session=False)
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/notifications/unread-count')
@login_required
def api_unread_count():
    """Get unread notification count."""
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    
    count = Notification.query.filter_by(student_id=student_id, is_read=False).count()
    return jsonify({'success': True, 'unread_count': count})

# ==================== ADMIN NOTIFICATION API ROUTES ====================

@app.route('/api/admin/notifications')
@src_required
def api_admin_get_notifications():
    """Get notifications for the logged-in admin."""
    admin_id = session.get('admin_id')
    if not admin_id:
        return jsonify({'success': False, 'error': 'Not logged in'}), 401
    
    # Check for unattended feedback periodically
    check_unattended_feedback()
    
    notifications = Notification.query.filter_by(recipient_type='admin').order_by(Notification.created_at.desc()).limit(30).all()
    unread_count = Notification.query.filter_by(recipient_type='admin', is_read=False).count()
    
    return jsonify({
        'success': True,
        'notifications': [n.to_dict() for n in notifications],
        'unread_count': unread_count
    })

@app.route('/api/admin/notifications/mark-read', methods=['POST'])
@src_required
def api_admin_mark_notifications_read():
    """Mark admin notifications as read."""
    data = request.get_json()
    notification_ids = data.get('ids', [])
    
    if notification_ids:
        Notification.query.filter(
            Notification.id.in_(notification_ids),
            Notification.recipient_type == 'admin'
        ).update({'is_read': True}, synchronize_session=False)
    else:
        # Mark all as read
        Notification.query.filter_by(recipient_type='admin', is_read=False).update({'is_read': True}, synchronize_session=False)
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/admin/notifications/unread-count')
@src_required
def api_admin_unread_count():
    """Get unread notification count for admin."""
    count = Notification.query.filter_by(recipient_type='admin', is_read=False).count()
    return jsonify({'success': True, 'unread_count': count})

# ==================== ACTIVE LEARNING API ROUTES ====================

@app.route('/api/admin/sentiment/record-correction', methods=['POST'])
@src_required
def api_record_sentiment_correction():
    """Record an admin correction to sentiment for active learning"""
    data = request.get_json()
    feedback_id = data.get('feedback_id')
    original = data.get('original_sentiment')
    corrected = data.get('corrected_sentiment')
    confidence = data.get('confidence', 0.0)

    if not feedback_id or not original or not corrected:
        return jsonify({'success': False, 'error': 'Missing required fields'}), 400

    try:
        from database import db, SentimentCorrection, Feedback
        
        # Record the correction
        correction = SentimentCorrection(
            feedback_id=feedback_id,
            original_sentiment=original,
            corrected_sentiment=corrected,
            admin_name=session.get('admin_name', 'Admin'),
            confidence_before=confidence,
        )
        db.session.add(correction)
        
        # Update the feedback record
        feedback = Feedback.query.get(feedback_id)
        if feedback:
            feedback.sentiment = corrected.lower()
            feedback.confidence_score = 100.0
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/sentiment/correction-stats')
@src_required
def api_sentiment_correction_stats():
    """Get sentiment correction statistics"""
    try:
        from database import SentimentCorrection
        total = SentimentCorrection.query.count()
        changed = SentimentCorrection.query.filter(
            SentimentCorrection.original_sentiment != SentimentCorrection.corrected_sentiment
        ).count()
        return jsonify({
            'success': True,
            'total_corrections': total,
            'changed': changed,
            'accuracy_improvement': round(changed / total * 100, 1) if total > 0 else 0,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/sentiment/analyze')
@src_required
def api_analyze_sentiment():
    """Enhanced sentiment analysis with aspects and emotions"""
    text = request.args.get('text', '')
    if not text:
        return jsonify({'success': False, 'error': 'No text provided'}), 400

    try:
        from sentiment_analyzer import process_feedback
        result = process_feedback(text)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/logs')
@src_required
def admin_logs():
    log_type = request.args.get('log_type', '')
    level = request.args.get('level', '')
    days = request.args.get('days', 7, type=int)
    page = request.args.get('page', 1, type=int)
    
    query = SystemLog.query
    if log_type:
        query = query.filter_by(log_type=log_type)
    if level:
        query = query.filter_by(level=level)
    if days:
        query = query.filter(SystemLog.timestamp >= datetime.utcnow() - timedelta(days=days))
    
    paginated = query.order_by(SystemLog.timestamp.desc()).paginate(page=page, per_page=25, error_out=False)
    logs = paginated.items
    
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Stats
    total_logs = SystemLog.query.count()
    today_logs = SystemLog.query.filter(SystemLog.timestamp >= today_start).count()
    warning_count = SystemLog.query.filter_by(level='WARNING').count()
    error_count = SystemLog.query.filter_by(level='ERROR').count()
    
    stats = {
        'total': total_logs,
        'today': today_logs,
        'warnings': warning_count,
        'errors': error_count
    }
    
    # Severity distribution for chart
    severity_distribution = {
        'INFO': SystemLog.query.filter_by(level='INFO').count(),
        'WARNING': SystemLog.query.filter_by(level='WARNING').count(),
        'ERROR': SystemLog.query.filter_by(level='ERROR').count()
    }
    
    # Log type distribution
    type_distribution = {}
    for t in ['auth', 'feedback', 'admin', 'system']:
        type_distribution[t] = SystemLog.query.filter_by(log_type=t).count()
    
    # Activity trend (last 7 days)
    log_trend = []
    for i in range(6, -1, -1):
        day = now - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        day_count = SystemLog.query.filter(
            SystemLog.timestamp >= day_start,
            SystemLog.timestamp < day_end
        ).count()
        log_trend.append({
            'date': day_start.strftime('%a'),
            'count': day_count
        })
    
    return render_template('admin_logs.html', logs=logs, pagination=paginated, stats=stats,
                         current_type=log_type, current_level=level, current_days=days,
                         severity_distribution=severity_distribution,
                         type_distribution=type_distribution,
                         log_trend=log_trend)

@app.route('/admin/export/excel')
@src_required
def export_excel():
    feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback Data"
    
    headers = ['ID', 'Date', 'Category', 'Original Feedback', 'Cleaned Text', 'Sentiment', 'Urgency', 'Status', 'Assigned To', 'SRC Response', 'Has Profanity']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row, f in enumerate(feedbacks, 2):
        ws.cell(row=row, column=1, value=f.id)
        ws.cell(row=row, column=2, value=f.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=3, value=f.category)
        ws.cell(row=row, column=4, value=f.feedback_text[:500])
        ws.cell(row=row, column=5, value=f.cleaned_text[:500] if f.cleaned_text else '')
        ws.cell(row=row, column=6, value=f.sentiment)
        ws.cell(row=row, column=7, value=f"{f.urgency_score}/5")
        ws.cell(row=row, column=8, value=f.status)
        ws.cell(row=row, column=9, value=f.assigned_to or '')
        ws.cell(row=row, column=10, value=f.src_response or '')
        ws.cell(row=row, column=11, value='Yes' if f.has_profanity else 'No')
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_admin_action(session['admin_name'], 'Export Excel', f'Exported {len(feedbacks)} feedback records')
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=feedback_export_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@app.route('/api/admin/export/students')
@src_required
def api_export_students():
    """Export student logs as JSON for CSV conversion"""
    students = Student.query.order_by(Student.created_at.desc()).all()
    return jsonify({
        'success': True,
        'students': [{
            'student_id': s.student_id,
            'full_name': s.full_name,
            'email': s.email,
            'department': s.department or '-',
            'year': s.year_of_study or '-',
            'registered_at': s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else '-',
            'last_login': s.last_login.strftime('%Y-%m-%d %H:%M') if s.last_login else '-',
            'status': 'Active' if s.is_active else 'Inactive'
        } for s in students]
    })

@app.route('/api/admin/export/analytics')
@src_required
def api_export_analytics():
    """Export advanced analytics as JSON for CSV conversion"""
    all_feedback = Feedback.query.all()
    total = len(all_feedback)
    resolved = sum(1 for f in all_feedback if f.status == 'resolved')
    pending = sum(1 for f in all_feedback if f.status == 'pending')
    in_progress = sum(1 for f in all_feedback if f.status == 'in_progress')
    positive = sum(1 for f in all_feedback if f.sentiment == 'positive')
    negative = sum(1 for f in all_feedback if f.sentiment == 'negative')
    neutral = sum(1 for f in all_feedback if f.sentiment == 'neutral')
    avg_urgency = sum(f.urgency_score for f in all_feedback) / total if total > 0 else 0
    resolution_rate = (resolved / total * 100) if total > 0 else 0

    metrics = [
        {'name': 'Total Feedback', 'value': total, 'description': 'Total number of feedback submissions'},
        {'name': 'Resolved', 'value': resolved, 'description': 'Feedback marked as resolved'},
        {'name': 'Pending', 'value': pending, 'description': 'Feedback awaiting review'},
        {'name': 'In Progress', 'value': in_progress, 'description': 'Feedback currently being processed'},
        {'name': 'Positive Sentiment', 'value': positive, 'description': 'Feedback with positive sentiment'},
        {'name': 'Negative Sentiment', 'value': negative, 'description': 'Feedback with negative sentiment'},
        {'name': 'Neutral Sentiment', 'value': neutral, 'description': 'Feedback with neutral sentiment'},
        {'name': 'Average Urgency', 'value': f'{avg_urgency:.1f}/5', 'description': 'Average urgency score across all feedback'},
        {'name': 'Resolution Rate', 'value': f'{resolution_rate:.1f}%', 'description': 'Percentage of feedback resolved'},
    ]

    return jsonify({'success': True, 'metrics': metrics})

@app.route('/api/admin/export/logs')
@src_required
def api_export_logs():
    """Export system logs as JSON for CSV conversion"""
    logs = SystemLog.query.order_by(SystemLog.timestamp.desc()).all()
    return jsonify({
        'success': True,
        'logs': [{
            'timestamp': l.timestamp.strftime('%Y-%m-%d %H:%M:%S') if l.timestamp else '-',
            'level': l.level or 'INFO',
            'user': l.user_id or l.user_type or 'System',
            'action': l.action or l.log_type or '-',
            'details': l.details or '-',
            'ip_address': l.ip_address or '-'
        } for l in logs]
    })

@app.route('/api/admin/trending')
@src_required
def api_trending_issues():
    """Get trending issues for admin dashboard"""
    trends = get_trending_issues()
    workload = get_department_workload()
    return jsonify({
        'success': True,
        'trends': trends,
        'workload': workload,
        'total_issues': sum(t['count'] for t in trends),
    })

@app.route('/api/chat/export/<int:room_id>')
@login_required
def api_chat_export(room_id):
    """Export all messages in a chat room as JSON for CSV conversion"""
    room = ChatRoom.query.get_or_404(room_id)
    messages = ChatMessage.query.filter_by(room_id=room_id).order_by(ChatMessage.created_at.asc()).all()
    return jsonify({
        'success': True,
        'room': {
            'id': room.id,
            'name': room.name,
            'category': room.category or 'General',
        },
        'messages': [{
            'timestamp': m.created_at.strftime('%Y-%m-%d %H:%M:%S') if m.created_at else '-',
            'sender': m.student.full_name if m.student else 'Unknown',
            'message': m.message or '',
            'message_type': m.message_type or 'text',
            'sentiment': m.sentiment or 'N/A',
        } for m in messages]
    })

@app.route('/api/admin/recommendation/feedback', methods=['POST'])
@src_required
def api_recommendation_feedback():
    """Record feedback on recommendation effectiveness"""
    data = request.get_json()
    category = data.get('category', '')
    solution_key = data.get('solution_key', '')
    was_helpful = data.get('was_helpful', False)
    resolved = data.get('resolved', False)
    engine = get_engine()
    engine.record_solution_feedback(category, solution_key, was_helpful, resolved)
    return jsonify({'success': True})

@app.route('/admin/analytics')
@src_required
def admin_analytics():
    """Advanced analytics dashboard with predictive insights"""
    all_feedback = Feedback.query.all()
    now = datetime.utcnow()
    
    # ==================== Predictive Analytics ====================
    from predictive_analytics import predict_events_combined

    # Get chat messages from the last 7 days for predictions
    seven_days_ago = now - timedelta(days=7)
    recent_chat_messages = ChatMessage.query.filter(ChatMessage.created_at >= seven_days_ago).all()
    
    # Get unresolved feedback older than 48 hours
    unresolved_feedback = Feedback.query.filter(
        Feedback.status != 'Resolved',
        Feedback.created_at <= now - timedelta(hours=48)
    ).all()
    
    prediction_result = predict_events_combined(
        room_messages=recent_chat_messages,
        feedback_items=unresolved_feedback,
        now=now
    )
    
    # ==================== Category Stats ====================
    category_stats = {}
    for f in all_feedback:
        if f.category not in category_stats:
            category_stats[f.category] = {
                'total': 0, 'positive': 0, 'negative': 0, 'neutral': 0,
                'avg_urgency': 0, 'resolution_rate': 0, 'resolved': 0
            }
        stats = category_stats[f.category]
        stats['total'] += 1
        stats[f.sentiment.lower()] += 1
        if f.status == 'Resolved':
            stats['resolved'] += 1
        stats['avg_urgency'] += f.urgency_score
    
    for cat in category_stats:
        s = category_stats[cat]
        if s['total'] > 0:
            s['avg_urgency'] = round(s['avg_urgency'] / s['total'], 1)
            s['resolution_rate'] = round((s['resolved'] / s['total']) * 100, 1)
            s['positive_pct'] = round((s['positive'] / s['total']) * 100, 1)
            s['negative_pct'] = round((s['negative'] / s['total']) * 100, 1)
    
    # ==================== Location Analysis ====================
    location_complaints = {}
    for f in all_feedback:
        if f.location:
            loc = f.location
            location_complaints[loc] = location_complaints.get(loc, 0) + 1
    top_locations = sorted(location_complaints.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # ==================== Time Analysis ====================
    hour_distribution = [0] * 24
    day_distribution = [0] * 7
    for f in all_feedback:
        hour_distribution[f.created_at.hour] += 1
        day_distribution[f.created_at.weekday()] += 1
    
    # ==================== Sentiment Trend (30 days) ====================
    sentiment_trend = []
    for i in range(30, -1, -1):
        day = now - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        day_feedback = [f for f in all_feedback if day_start <= f.created_at < day_end]
        sentiment_trend.append({
            'date': day_start.strftime('%b %d'),
            'positive': sum(1 for f in day_feedback if f.sentiment == 'Positive'),
            'negative': sum(1 for f in day_feedback if f.sentiment == 'Negative'),
            'neutral': sum(1 for f in day_feedback if f.sentiment == 'Neutral'),
            'total': len(day_feedback)
        })
    
    # ==================== Department Performance ====================
    dept_stats = {}
    for f in all_feedback:
        dept = f.responsible_department or 'Unassigned'
        if dept not in dept_stats:
            dept_stats[dept] = {'total': 0, 'resolved': 0, 'avg_time': 0, 'total_time': 0}
        dept_stats[dept]['total'] += 1
        if f.status == 'Resolved' and f.resolved_at:
            dept_stats[dept]['resolved'] += 1
            time_taken = (f.resolved_at - f.created_at).total_seconds() / 3600
            dept_stats[dept]['total_time'] += time_taken
    
    dept_performance = []
    for dept, s in dept_stats.items():
        dept_performance.append({
            'name': dept,
            'total': s['total'],
            'resolved': s['resolved'],
            'resolution_rate': round((s['resolved'] / s['total']) * 100, 1) if s['total'] > 0 else 0,
            'avg_time_hours': round(s['total_time'] / s['resolved'], 1) if s['resolved'] > 0 else 0
        })
    dept_performance.sort(key=lambda x: x['total'], reverse=True)
    
    # ==================== Recommendation Insights ====================
    learner = RecommendationLearner()
    solution_templates = SolutionTemplate.query.all()
    learning_report = learner.generate_report(all_feedback, solution_templates)
    
    # ==================== Overall Metrics ====================
    total = len(all_feedback)
    resolved_count = sum(1 for f in all_feedback if f.status == 'Resolved')
    pending_count = sum(1 for f in all_feedback if f.status == 'Pending')
    in_progress_count = sum(1 for f in all_feedback if f.status == 'In Progress')
    urgent_count = sum(1 for f in all_feedback if f.urgency_score >= 4)
    
    # Resolution time (average in hours)
    resolved_feedback = [f for f in all_feedback if f.status == 'Resolved' and f.resolved_at]
    avg_resolution_time = 0
    if resolved_feedback:
        times = [(f.resolved_at - f.created_at).total_seconds() / 3600 for f in resolved_feedback]
        avg_resolution_time = round(sum(times) / len(times), 1)
    
    return render_template('admin_analytics.html',
                         category_stats=category_stats,
                         top_locations=top_locations,
                         hour_distribution=hour_distribution,
                         day_distribution=day_distribution,
                         sentiment_trend=sentiment_trend,
                         dept_performance=dept_performance,
                         predictions=prediction_result.get('predictions', []),
                         early_warning_level=prediction_result.get('early_warning_level', 'N/A'),
                         max_confidence=prediction_result.get('max_confidence', 0),
                         total=total,
                         resolved_count=resolved_count,
                         pending_count=pending_count,
                         in_progress_count=in_progress_count,
                         urgent_count=urgent_count,
                         avg_resolution_time=avg_resolution_time,
                         learning_report=learning_report)

@app.route('/admin/chat')
@app.route('/admin/chat/rooms')
@src_required
def admin_chat_rooms():
    """Unified admin chat center.

    Rooms are the primary objects shown to admins. Messages are loaded per room
    so the dashboard does not need a separate 'Chat Messages' page.
    """
    rooms = ChatRoom.query.options(
        db.joinedload(ChatRoom.messages).joinedload(ChatMessage.student)
    ).order_by(ChatRoom.last_activity.desc(), ChatRoom.created_at.desc()).all()

    room_cards = []
    total_messages = 0
    total_flagged = 0
    active_rooms = 0

    for room in rooms:
        messages = sorted(
            list(room.messages),
            key=lambda m: m.created_at or datetime.min
        )
        flagged = [m for m in messages if m.is_flagged]
        sentiment_counts = {
            'positive': sum(1 for m in messages if m.sentiment == 'Positive'),
            'negative': sum(1 for m in messages if m.sentiment == 'Negative'),
            'neutral': sum(1 for m in messages if m.sentiment not in ('Positive', 'Negative'))
        }
        last_message = messages[-1] if messages else None

        room_cards.append({
            'room': room,
            'messages': [m.to_dict() for m in messages],
            'message_count': len(messages),
            'member_count': ChatRoomMember.query.filter_by(room_id=room.id).count(),
            'flagged_count': len(flagged),
            'sentiment_counts': sentiment_counts,
            'last_message': last_message.to_dict() if last_message else None,
        })
        total_messages += len(messages)
        total_flagged += len(flagged)
        active_rooms += 1 if room.is_active else 0

    return render_template(
        'admin_chat_rooms.html',
        rooms=room_cards,
        total_rooms=len(rooms),
        active_rooms=active_rooms,
        total_messages_all=total_messages,
        total_flagged_all=total_flagged,
        unified_chat=True
    )


@app.route('/admin/chat/rooms/<int:room_id>/delete', methods=['POST'])
@src_required
def admin_delete_chat_room(room_id):
    room = ChatRoom.query.get_or_404(room_id)
    ChatMessage.query.filter_by(room_id=room_id).delete()
    ChatRoomMember.query.filter_by(room_id=room_id).delete()
    ChatRoomSentiment.query.filter_by(room_id=room_id).delete()
    db.session.delete(room)
    db.session.commit()
    log_admin_action(session['admin_name'], 'Delete Chat Room', f'Deleted room ID: {room_id}')
    return redirect(url_for('admin_chat_rooms'))


@app.route('/admin/chat/messages')
@src_required
def admin_chat_messages():
    """Legacy URL retained so old bookmarks do not break.

    The admin now uses the unified chat center where messages are expanded
    inside each room card.
    """
    return redirect(url_for('admin_chat_rooms'))


@app.route('/admin/change-password', methods=['GET', 'POST'])
@src_required
@limiter.limit("5 per hour")
def admin_change_password():
    admin_user = SRCUser.query.get(session['admin_id'])

    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not admin_user or not check_password_hash(admin_user.password_hash, current_password):
            return render_template('change_password.html', role='admin', error='Current password is incorrect')

        if new_password != confirm_password:
            return render_template('change_password.html', role='admin', error='New passwords do not match')

        if len(new_password) < 6:
            return render_template('change_password.html', role='admin', error='New password must be at least 6 characters')

        if check_password_hash(admin_user.password_hash, new_password):
            return render_template('change_password.html', role='admin', error='New password must be different from your current password')

        admin_user.password_hash = generate_password_hash(new_password)
        db.session.commit()

        log_admin_action(admin_user.full_name, 'Password Change', 'Password changed successfully from dashboard')
        add_db_log('auth', 'INFO', 'admin', admin_user.username, 'Password Change', '')
        return render_template('change_password.html', role='admin', success='Your password has been updated successfully.')

    return render_template('change_password.html', role='admin')

@app.route('/admin/logout', methods=['POST'])
@src_required
def admin_logout():
    admin_name = session.get('admin_name')
    log_admin_action(admin_name, 'Logout', 'Admin logged out')
    session.clear()
    return redirect(url_for('index'))

# ==================== PUBLIC ROUTES ====================

@app.route('/public')
def public_board():
    all_feedback = Feedback.query.all()

    sentiment_counts = {
        'positive': sum(1 for f in all_feedback if f.sentiment == 'Positive'),
        'negative': sum(1 for f in all_feedback if f.sentiment == 'Negative'),
        'neutral': sum(1 for f in all_feedback if f.sentiment == 'Neutral')
    }

    category_sentiment = {}
    for f in all_feedback:
        if f.category not in category_sentiment:
            category_sentiment[f.category] = {'positive': 0, 'negative': 0, 'neutral': 0, 'total': 0}
        category_sentiment[f.category][f.sentiment.lower()] += 1
        category_sentiment[f.category]['total'] += 1

    top_complaints = Feedback.query.filter(Feedback.sentiment == 'Negative').order_by(Feedback.urgency_score.desc()).limit(5).all()
    forum_summary = get_forum_sentiment_summary(ForumTopic.query.all())

    # ==================== SRC Performance ====================
    # Engagement increase = compare last 7 days vs previous 7 days.
    # Engagement definition: (Feedback submissions + Forum replies).
    now = datetime.utcnow()
    current_start = (now - timedelta(days=7))
    previous_start = (now - timedelta(days=14))

    current_feedback_count = Feedback.query.filter(Feedback.created_at >= current_start).count()
    previous_feedback_count = Feedback.query.filter(Feedback.created_at >= previous_start, Feedback.created_at < current_start).count()

    current_replies_count = ForumReply.query.filter(ForumReply.created_at >= current_start).count()
    previous_replies_count = ForumReply.query.filter(ForumReply.created_at >= previous_start, ForumReply.created_at < current_start).count()

    current_engagement = current_feedback_count + current_replies_count
    previous_engagement = previous_feedback_count + previous_replies_count

    if previous_engagement > 0:
        engagement_increase_pct = ((current_engagement - previous_engagement) / previous_engagement) * 100
    else:
        # If there was no engagement in the previous period, treat any engagement as "new".
        engagement_increase_pct = 100.0 if current_engagement > 0 else 0.0

    engagement_increase_pct = round(engagement_increase_pct, 1)
    if engagement_increase_pct > 0:
        engagement_arrow = '↑'
    elif engagement_increase_pct < 0:
        engagement_arrow = '↓'
    else:
        engagement_arrow = '→'

    positive_rate = (sentiment_counts['positive'] / len(all_feedback) * 100) if len(all_feedback) > 0 else 0

    return render_template(
        'public_board.html',
        sentiment_counts=sentiment_counts,
        category_sentiment=category_sentiment,
        top_complaints=top_complaints,
        total_feedback=len(all_feedback),
        forum_summary=forum_summary,
        src_positive_rate=round(positive_rate, 0),
        engagement_increase_pct=abs(engagement_increase_pct),
        engagement_arrow=engagement_arrow
    )

@app.route('/')
def index():
    # Live stats for landing page
    all_feedback = Feedback.query.all()
    total_feedback = len(all_feedback)
    total_students = Student.query.count()
    resolved_count = sum(1 for f in all_feedback if f.status == 'Resolved')
    positive_count = sum(1 for f in all_feedback if f.sentiment == 'Positive')
    positive_rate = round((positive_count / total_feedback * 100) if total_feedback > 0 else 0, 0)

    # Recent announcements
    now = datetime.utcnow()
    recent_announcements = Announcement.query.filter(
        Announcement.is_active == True,
        (Announcement.expires_at.is_(None) | (Announcement.expires_at > now))
    ).order_by(Announcement.created_at.desc()).limit(3).all()

    # Engagement stats
    week_ago = now - timedelta(days=7)
    recent_feedback = sum(1 for f in all_feedback if f.created_at >= week_ago)
    recent_replies = ForumReply.query.filter(ForumReply.created_at >= week_ago).count()

    return render_template('index.html',
        total_feedback=total_feedback,
        total_students=total_students,
        resolved_count=resolved_count,
        positive_rate=positive_rate,
        recent_announcements=recent_announcements,
        recent_engagement=recent_feedback + recent_replies
    )

@app.route('/api/set-theme', methods=['POST'])
@student_required
def set_theme():
    theme = request.json.get('theme')
    if theme not in ('light', 'dark'):
        return jsonify({'error': 'Invalid theme selection'}), 400

    student = Student.query.filter_by(student_id=session['student_id']).first()
    if student:
        student.theme_preference = theme
        db.session.commit()
    return jsonify({'success': True})

if __name__ == '__main__':
    log_system_action('System', 'Startup', 'HTU SRC Feedback System started with all features')
    debug_mode = os.environ.get('FLASK_DEBUG', '0').lower() in ('1', 'true', 'yes')
    socketio.run(app, debug=debug_mode)
